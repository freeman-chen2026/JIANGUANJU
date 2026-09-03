import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta, date
import tempfile
import os
import re
import json
import pdfplumber
from collections import defaultdict
import traceback

# ==============================
# 通用工具函数（供各功能使用）
# ==============================
def parse_duration(dur_str) -> int:
    if pd.isna(dur_str):
        return 0
    s = str(dur_str).strip().replace('：', ':')
    s_lower = s.lower()
    days_pattern = re.compile(r'(\d+)\s*days?\s*,?\s*(\d+):(\d{2})(?::(\d{2}))?', re.IGNORECASE)
    match = days_pattern.search(s_lower)
    if match:
        days = int(match.group(1))
        hours = int(match.group(2))
        minutes = int(match.group(3))
        return days * 24 * 60 + hours * 60 + minutes
    time_pattern = re.compile(r'(\d+):(\d{2})(?::(\d{2}))?')
    match = time_pattern.search(s)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        return hours * 60 + minutes
    try:
        return int(float(s))
    except:
        return 0

def format_duration(total_minutes: int) -> str:
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours}:{minutes:02d}:00"

def detect_header_row(df_raw, keywords):
    best_score = -1
    best_row = 0
    for i in range(min(10, len(df_raw))):
        row_values = [str(v).strip() for v in df_raw.iloc[i].values]
        score = sum(1 for kw in keywords if any(kw in val for val in row_values))
        if score > best_score:
            best_score = score
            best_row = i
    if best_score > 0:
        return best_row, True
    else:
        return 0, False

def read_excel_with_auto_header(file, keywords):
    raw = pd.read_excel(file, header=None, dtype=str)
    header_idx, found = detect_header_row(raw, keywords)
    if not found:
        st.warning("未检测到表头，将使用第一行作为列名，可能导致错误。")
        return pd.read_excel(file, header=0)
    df = pd.read_excel(file, header=header_idx)
    df = df.dropna(axis=1, how='all')
    df = df.loc[:, ~df.columns.str.contains('Unnamed', case=False)]
    df = df.reset_index(drop=True)
    return df

def auto_match_column(df, candidates):
    cols_lower = {col.strip().lower(): col for col in df.columns}
    for cand in candidates:
        cand_lower = cand.strip().lower()
        if cand_lower in cols_lower:
            return cols_lower[cand_lower]
    for cand in candidates:
        for col in df.columns:
            if cand in col:
                return col
    return None

def format_time(value):
    if pd.isna(value) or value == "" or value is None:
        return ""
    if isinstance(value, str):
        if ":" in value:
            parts = value.split(":")
            if len(parts) == 2:
                return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}:00"
            elif len(parts) == 3:
                return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}:{parts[2].zfill(2)}"
        return value
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M:%S")
    return str(value)

# ==============================
# 功能 A：每日飞行数据自动更新
# ==============================
def update_excel1(excel1_path, excel2_df, flight_col, dep_col, arr_col, reg_col, purpose_col):
    wb = load_workbook(excel1_path)
    ws = wb.active

    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    ws.cell(row=2, column=10).value = f"*昨日总飞行时间\n（昨日指{yesterday.month}月{yesterday.day}日）*"

    valid_mask = excel2_df[flight_col].notna()
    valid_df = excel2_df[valid_mask].copy()

    total_minutes = 0
    for val in valid_df[flight_col]:
        total_minutes += parse_duration(val)

    ws.cell(row=3, column=10).value = format_duration(total_minutes)  # J3
    ws.cell(row=3, column=11).value = len(valid_df)                  # K3

    old_l3 = ws.cell(row=3, column=12).value
    old_minutes = parse_duration(old_l3) if old_l3 is not None else 0
    new_total = old_minutes + total_minutes
    ws.cell(row=3, column=12).value = format_duration(new_total)     # L3

    # ---------- 修改部分：M3 同时判断调机飞行和公务飞行 ----------
    has_diaoji = False
    has_gongwu = False
    if purpose_col and not valid_df.empty:
        for val in valid_df[purpose_col].dropna():
            purpose_str = str(val).strip()
            if "调机" in purpose_str or "维修" in purpose_str:
                has_diaoji = True
            else:
                # 其他用途视为公务飞行（载客、共享租赁等）
                has_gongwu = True

    # 组合M3内容
    m3_parts = []
    if has_gongwu:
        m3_parts.append("公务飞行")
    if has_diaoji:
        m3_parts.append("调机飞行")
    # 如果没有任何用途（极少情况），默认填公务飞行
    m3_value = "、".join(m3_parts) if m3_parts else "公务飞行"
    ws.cell(row=3, column=13).value = m3_value                      # M3
    # ---------- 修改结束 ----------

    # N3：收集所有出发城市和到达城市，去重后顿号连接
    locations = set()
    for _, row in valid_df.iterrows():
        dep = str(row[dep_col]).strip() if pd.notna(row[dep_col]) else ''
        arr = str(row[arr_col]).strip() if pd.notna(row[arr_col]) else ''
        if dep:
            locations.add(dep)
        if arr:
            locations.add(arr)
    location_list = sorted(locations)
    ws.cell(row=3, column=14).value = '、'.join(location_list)        # N3

    reg_series = valid_df[reg_col].dropna()
    ws.cell(row=3, column=15).value = len(reg_series.astype(str).unique())  # O3

    stats = {
        '昨日飞行时间': format_duration(total_minutes),
        '架次': len(valid_df),
        '注册号数量': len(reg_series.astype(str).unique()),
        '截止昨日总飞行时间': format_duration(old_minutes),
        '截止今日总飞行时间': format_duration(new_total),
    }

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    return tmp.name, stats

def run_feature_a():
    excel1_file = st.file_uploader("📂 上传：昨日飞行数据（operation-每日飞行数据）", type=["xlsx", "xlsm"], key="a_excel1")
    excel2_file = st.file_uploader("📂 上传：航段数据导出（昨日B机）", type=["xlsx", "xlsm"], key="a_excel2")

    if excel1_file and excel2_file:
        with st.spinner("正在自动处理..."):
            try:
                keywords = ["客户", "航班号", "出发城市", "到达城市", "实际飞行时间", "飞机注册号", "用途"]
                excel2_df = read_excel_with_auto_header(excel2_file, keywords)
                if excel2_df.empty or len(excel2_df.columns) == 0:
                    st.error("航段数据没有有效列，请检查文件格式。")
                    return

                flight_col = auto_match_column(excel2_df, ["实际飞行时间", "飞行时间", "航段时间"])
                dep_col = auto_match_column(excel2_df, ["出发城市", "起飞机场", "出发地"])
                arr_col = auto_match_column(excel2_df, ["到达城市", "目的地机场", "到达地"])
                reg_col = auto_match_column(excel2_df, ["飞机注册号", "注册号", "机号"])
                purpose_col = auto_match_column(excel2_df, ["用途", "任务性质", "飞行任务"])

                missing = []
                if not flight_col: missing.append("飞行时间")
                if not dep_col: missing.append("出发城市")
                if not arr_col: missing.append("到达城市")
                if not reg_col: missing.append("飞机注册号")
                if not purpose_col: missing.append("用途")
                if missing:
                    st.error(f"未能自动匹配以下列：{', '.join(missing)}，请检查文件列名是否包含关键词。")
                    return

                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp1:
                    tmp1.write(excel1_file.getvalue())
                    excel1_path = tmp1.name

                output_path, stats = update_excel1(
                    excel1_path, excel2_df, flight_col, dep_col, arr_col, reg_col, purpose_col
                )

                st.success("✅ 处理完成")
                col1, col2, col3 = st.columns(3)
                col1.metric("昨日总飞行时间", stats['昨日飞行时间'])
                col2.metric("架次", stats['架次'])
                col3.metric("使用航空器数量", stats['注册号数量'])

                col4, col5 = st.columns(2)
                col4.metric("截止昨日总飞行时间", stats['截止昨日总飞行时间'])
                col5.metric("截止今日总飞行时间", stats['截止今日总飞行时间'])

                yesterday = datetime.now().date() - timedelta(days=1)
                file_name = f"中南-深圳局-天成商务航空有限公司-{yesterday.month}月{yesterday.day}日飞行数据.xlsx"

                with open(output_path, 'rb') as f:
                    st.download_button(
                        label="📥 下载更新后的文件",
                        data=f.read(),
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                os.unlink(excel1_path)
                os.unlink(output_path)

            except Exception as e:
                st.error(f"处理出错：{e}")
                if 'excel1_path' in locals() and os.path.exists(excel1_path):
                    os.unlink(excel1_path)
                if 'output_path' in locals() and os.path.exists(output_path):
                    os.unlink(output_path)

# ==============================
# 功能 B：模板生成备案表
# ==============================
DEFAULT_ICAO_MAP = {
    "B65AP": "GLF4",
    "B652R": "GLF4",
    "B652S": "GLF4",
    "B8105": "GLEX",
    "B8309": "GLF5",
    "B652Q": "GLF4",
    "B3926": "LJ60",
    "B8160": "GLF5",
    "B8262": "GLF4",
    "B8292": "GLF5",
}

def find_header_row(df_raw):
    keywords = ["客户", "航班号", "飞机注册号", "出发日期"]
    for idx, row in df_raw.iterrows():
        row_str = " ".join([str(v) for v in row.values if pd.notna(v)])
        if all(kw in row_str for kw in keywords):
            return idx
    return 0

def parse_uploaded_file(uploaded_file):
    df_all = pd.read_excel(uploaded_file, sheet_name=0, header=None)
    header_row = find_header_row(df_all)
    df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row)
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    return df

def map_usage(usage):
    usage = str(usage).strip()
    if "调机" in usage:
        return "调机飞行", "调机飞行"
    elif "载客" in usage or "共享租赁" in usage:
        return "公务飞行", "私用飞行"
    else:
        return "公务飞行", "私用飞行"

def parse_duration_to_minutes(duration_str):
    if pd.isna(duration_str) or duration_str == "" or duration_str is None:
        return None
    s = str(duration_str).strip()
    try:
        minutes = parse_duration(s)
        return minutes
    except:
        return None

def combine_date_time(date_val, time_val):
    if pd.isna(date_val) or pd.isna(time_val):
        return None
    try:
        date_str = str(date_val).strip()
        time_str = str(time_val).strip()
        if time_str.count(":") == 2:
            dt_str = f"{date_str} {time_str}"
        else:
            dt_str = f"{date_str} {time_str}:00"
        return pd.to_datetime(dt_str)
    except:
        return None

def run_feature_b():
    default_supervision = "深圳局"
    default_operator = "天成商务航空有限公司"
    icao_map = DEFAULT_ICAO_MAP.copy()

    if "template_wb" not in st.session_state:
        st.session_state.template_wb = None
    if "header_row" not in st.session_state:
        st.session_state.header_row = None
    if "data_start_row" not in st.session_state:
        st.session_state.data_start_row = None
    if "group_rows" not in st.session_state:
        st.session_state.group_rows = 4

    st.subheader("📂 模板管理")
    if st.session_state.template_wb is None:
        st.info("首次使用请上传模板文件。")
        template_file = st.file_uploader("上传：桌面-Jetops申请一览-每日通航运行情况跟踪表（天成商务航空有限公司）20260708", type=["xlsx"], key="b_template_upload")
        if template_file:
            try:
                wb = load_workbook(template_file)
                ws = wb.active
                header_row = None
                for row in range(1, 5):
                    if any("所属监管局" in str(cell.value) for cell in ws[row]):
                        header_row = row
                        break
                if header_row is None:
                    st.error("模板中未找到表头行，请确认模板有“所属监管局”字段。")
                    st.stop()
                data_start_row = header_row + 1
                st.session_state.template_wb = wb
                st.session_state.header_row = header_row
                st.session_state.data_start_row = data_start_row
                st.success("✅ 模板加载成功！现在可以上传数据文件。")
            except Exception as e:
                st.error(f"模板加载失败：{e}")
    else:
        st.success("✅ 模板已加载（如需更换，请点击下方按钮重置）")
        if st.button("重新上传模板", key="b_reset_template"):
            st.session_state.template_wb = None
            st.session_state.header_row = None
            st.session_state.data_start_row = None
            st.rerun()

    st.subheader("📊 数据上传")
    data_file = st.file_uploader("上传：航段数据导出（当日B机）", type=["xlsx"], key="b_data_upload")

    if data_file and st.session_state.template_wb is not None:
        try:
            df_raw = parse_uploaded_file(data_file)
            st.success(f"✅ 成功读取 {len(df_raw)} 条航段记录")
            if len(df_raw) > 20:
                st.warning(f"数据条数（{len(df_raw)}）超过模板预设的20行，多余数据将被忽略。")

            total_flights = len(df_raw)
            status_col = "航段状态" if "航段状态" in df_raw.columns else None
            actual_depart_col = "实际出发" if "实际出发" in df_raw.columns else None
            reg_col = "飞机注册号" if "飞机注册号" in df_raw.columns else None

            if status_col:
                landed = df_raw[df_raw[status_col].astype(str).str.contains("已执飞|已完成", na=False)]
                not_landed = df_raw[~df_raw[status_col].astype(str).str.contains("已执飞|已完成", na=False)]
            else:
                landed = pd.DataFrame()
                not_landed = df_raw

            if actual_depart_col:
                unlanded = not_landed[not_landed[actual_depart_col].notna()]
                not_executed = not_landed[not_landed[actual_depart_col].isna()]
            else:
                unlanded = pd.DataFrame()
                not_executed = not_landed

            landed_count = len(landed)
            unlanded_count = len(unlanded)
            not_executed_count = len(not_executed)

            if reg_col:
                all_regs = df_raw[reg_col].dropna().astype(str).unique()
                reg_list = sorted(all_regs)
            else:
                reg_list = []

            parts = []
            if landed_count > 0:
                parts.append(f"{landed_count}班已落地")
            if unlanded_count > 0:
                parts.append(f"{unlanded_count}班未落地")
            if not_executed_count > 0:
                parts.append(f"{not_executed_count}班未起飞")

            if not parts:
                report = "今天无飞行计划"
            else:
                if landed_count > 0 and unlanded_count == 0 and not_executed_count == 0:
                    report = "今天飞完了"
                else:
                    report = "今天" + "、".join(parts)

            if reg_list:
                plane_text = f"今日有飞行计划的飞机：{'、'.join(reg_list)}"
            else:
                plane_text = ""

            st.subheader("📊 飞行计划统计")
            col1, col2, col3 = st.columns(3)
            col1.metric("飞行计划总数", total_flights)
            col2.metric("已落地班次", landed_count)
            col3.metric("未落地班次", unlanded_count)
            col4, col5 = st.columns(2)
            col4.metric("未执行班次", not_executed_count)
            col5.metric("涉及飞机数", len(reg_list))
            if reg_list:
                st.write("**飞机注册号列表：**", "、".join(reg_list))

            full_report = report
            if plane_text:
                full_report += "\n" + plane_text
            st.text_area("📋 汇报文案（可复制）", full_report, height=120)

            wb = st.session_state.template_wb
            ws = wb.active
            header_row = st.session_state.header_row
            data_start_row = st.session_state.data_start_row
            group_rows = st.session_state.group_rows

            has_actual_depart = "实际出发" in df_raw.columns
            has_estimated_flight_time = "预计飞行时间" in df_raw.columns
            if not has_estimated_flight_time:
                st.warning("数据中没有“预计飞行时间”列，将使用原“预计到达”作为K列值。")

            records = []
            for idx, row in df_raw.iterrows():
                if idx >= 20:
                    break

                dt = pd.to_datetime(row["出发日期"])
                flight_date = f"{dt.year}/{dt.month}/{dt.day}"

                dep_city = str(row.get("出发城市", "")).strip()
                arr_city = str(row.get("到达城市", "")).strip()
                route = f"{dep_city}-{arr_city}" if dep_city and arr_city else f"{row['出发地']}-{row['到达地']}"

                run_type, oper_type = map_usage(row["用途"])

                if has_actual_depart:
                    actual_depart = format_time(row.get("实际出发", ""))
                    if actual_depart:
                        start_time = actual_depart
                    else:
                        start_time = format_time(row.get("计划出发", ""))
                else:
                    start_time = format_time(row.get("计划出发", ""))

                estimated_landing = ""
                if has_estimated_flight_time:
                    duration_str = row.get("预计飞行时间", "")
                    minutes = parse_duration_to_minutes(duration_str)
                    if minutes is not None:
                        date_val = row["出发日期"]
                        if has_actual_depart and pd.notna(row.get("实际出发")):
                            time_val = row["实际出发"]
                        else:
                            time_val = row["计划出发"]
                        if pd.notna(date_val) and pd.notna(time_val):
                            dt_obj = combine_date_time(date_val, time_val)
                            if dt_obj is not None:
                                new_dt = dt_obj + timedelta(minutes=minutes)
                                estimated_landing = new_dt.strftime("%H:%M:%S")
                    if not estimated_landing and "预计到达" in df_raw.columns:
                        estimated_landing = format_time(row.get("预计到达", ""))
                else:
                    if "预计到达" in df_raw.columns:
                        estimated_landing = format_time(row.get("预计到达", ""))

                actual_end = format_time(row.get("实际到达", "")) if "实际到达" in df_raw.columns else ""

                status = str(row.get("航段状态", "")).strip()
                is_landed = "是" if status in ["已执飞", "已完成"] else "否"

                reg = str(row["飞机注册号"]).strip().upper()
                icao_type = icao_map.get(reg, "")

                record = {
                    "A": default_supervision,
                    "B": default_operator,
                    "C": flight_date,
                    "D": run_type,
                    "E": oper_type,
                    "F": icao_type,
                    "G": reg,
                    "J": start_time,
                    "K": estimated_landing,
                    "L": is_landed,
                    "M": actual_end if is_landed == "是" else "",
                    "N": route,
                }
                records.append(record)

            for i, rec in enumerate(records):
                row_num = data_start_row + i * group_rows
                ws[f"A{row_num}"] = rec["A"]
                ws[f"B{row_num}"] = rec["B"]
                ws[f"C{row_num}"] = rec["C"]
                ws[f"D{row_num}"] = rec["D"]
                ws[f"E{row_num}"] = rec["E"]
                ws[f"F{row_num}"] = rec["F"]
                ws[f"G{row_num}"] = rec["G"]
                ws[f"J{row_num}"] = rec["J"]
                ws[f"K{row_num}"] = rec["K"]
                ws[f"L{row_num}"] = rec["L"]
                ws[f"M{row_num}"] = rec["M"]
                ws[f"N{row_num}"] = rec["N"]

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="⬇️ 下载填入数据的模板文件",
                data=output,
                file_name="每日通航运行情况跟踪表_生成.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="b_download"
            )

        except Exception as e:
            st.error(f"❌ 处理失败：{e}")
            st.exception(e)
    else:
        if st.session_state.template_wb is None:
            st.info("👆 请先上传模板。")
        else:
            st.info("👆 请上传航段数据 Excel 文件。")

# ==============================
# 功能 C：生成每日运行跟踪表
# ==============================
def run_feature_c():
    ICAO_MAP = {
        "B65AP": "GLF4",
        "B652R": "GLF4",
        "B652S": "GLF4",
        "B8105": "GLEX",
        "B8309": "GLF5",
        "B652Q": "GLF4",
        "B3926": "LJ60",
        "B8160": "GLF5",
        "B8262": "GLF4",
        "B8292": "GLF5",
    }

    st.markdown("上传 **天成商务航空每日运行跟踪** 模板和 **航段数据导出**，自动在模板最下方新增一行今日汇总数据。")

    template_file = st.file_uploader("📂 上传：Operation-每日通航运行情况-天成商务航空每日运行跟踪", type=["xlsx"], key="c_template")
    data_file = st.file_uploader("📂 上传：航段数据导出（当日B机）", type=["xlsx"], key="c_data")

    if template_file and data_file:
        with st.spinner("正在处理..."):
            try:
                keywords = ["客户", "航班号", "出发城市", "到达城市", "实际出发", "计划出发", "预计到达", "实际到达", "航段状态", "飞机注册号", "用途"]
                df = read_excel_with_auto_header(data_file, keywords)
                if df.empty:
                    st.error("航段数据为空或格式不正确。")
                    return

                required_cols = ["飞机注册号", "用途", "出发城市", "到达城市", "航段状态"]
                for col in required_cols:
                    if col not in df.columns:
                        st.error(f"航段数据缺少必要列：{col}")
                        return

                today = datetime.now().date()
                today_str = today.strftime("%Y/%m/%d")

                total_flights = len(df)
                actual_flights = total_flights

                usage_set = set()
                for u in df["用途"].dropna():
                    u_str = str(u).strip()
                    if "调机" in u_str:
                        usage_set.add("调机飞行")
                    else:
                        usage_set.add("公务飞行")
                run_types = "、".join(sorted(usage_set)) if usage_set else "公务飞行"

                status_col = "航段状态"
                has_actual_depart = "实际出发" in df.columns
                all_landed = all(str(s).strip() in ["已执飞", "已完成"] for s in df[status_col].dropna())
                if has_actual_depart:
                    any_started = any(pd.notna(s) for s in df["实际出发"])
                else:
                    any_started = any(str(s).strip() in ["已执飞", "已完成", "执飞中"] for s in df[status_col].dropna())

                if all_landed:
                    status_text = "已实施，全部结束"
                elif any_started:
                    status_text = "已实施，未全部结束"
                else:
                    status_text = "未实施"

                if has_actual_depart:
                    valid_depart = df[df["实际出发"].notna()]
                    if not valid_depart.empty:
                        times = valid_depart["实际出发"].apply(lambda x: str(x).strip())
                        earliest = min(times, key=lambda t: t if t else "99:99")
                        start_time = format_time(earliest)
                    else:
                        if "计划出发" in df.columns:
                            plan_times = df["计划出发"].dropna().apply(lambda x: str(x).strip())
                            if not plan_times.empty:
                                earliest_plan = min(plan_times, key=lambda t: t if t else "99:99")
                                start_time = format_time(earliest_plan)
                            else:
                                start_time = ""
                        else:
                            start_time = ""
                else:
                    if "计划出发" in df.columns:
                        plan_times = df["计划出发"].dropna().apply(lambda x: str(x).strip())
                        if not plan_times.empty:
                            earliest_plan = min(plan_times, key=lambda t: t if t else "99:99")
                            start_time = format_time(earliest_plan)
                        else:
                            start_time = ""
                    else:
                        start_time = ""

                if "预计到达" in df.columns:
                    plan_end_times = df["预计到达"].dropna().apply(lambda x: str(x).strip())
                    if not plan_end_times.empty:
                        latest_plan = max(plan_end_times, key=lambda t: t if t else "00:00")
                        plan_end = format_time(latest_plan)
                    else:
                        plan_end = ""
                else:
                    plan_end = ""

                if all_landed and "实际到达" in df.columns:
                    actual_end_times = df["实际到达"].dropna().apply(lambda x: str(x).strip())
                    if not actual_end_times.empty:
                        latest_actual = max(actual_end_times, key=lambda t: t if t else "00:00")
                        actual_end = format_time(latest_actual)
                    else:
                        actual_end = ""
                else:
                    actual_end = ""

                regs = df["飞机注册号"].dropna().astype(str).str.upper().unique()
                models = set()
                for reg in regs:
                    model = ICAO_MAP.get(reg, "")
                    if model:
                        models.add(model)
                model_str = "、".join(sorted(models)) if models else ""

                routes = []
                for _, row in df.iterrows():
                    dep = str(row.get("出发城市", "")).strip()
                    arr = str(row.get("到达城市", "")).strip()
                    if dep and arr:
                        routes.append(f"{dep}-{arr}")
                    elif dep:
                        routes.append(dep)
                    elif arr:
                        routes.append(arr)
                route_str = "、".join(routes)
                if len(route_str) > 100:
                    route_str = route_str[:100] + "..."

                supervision = "深圳局"
                category = "公务航空飞行"
                operator = "天成商务航空有限公司"
                yes = "是"

                wb = load_workbook(template_file)
                ws = wb.active

                target_row = None
                for row in range(2, ws.max_row + 2):
                    if ws.cell(row, 1).value is None or str(ws.cell(row, 1).value).strip() == "":
                        target_row = row
                        break
                if target_row is None:
                    target_row = ws.max_row + 1

                ws.cell(target_row, 1).value = supervision
                ws.cell(target_row, 2).value = today_str
                ws.cell(target_row, 3).value = category
                ws.cell(target_row, 4).value = operator
                ws.cell(target_row, 5).value = run_types
                ws.cell(target_row, 6).value = total_flights
                ws.cell(target_row, 7).value = actual_flights
                ws.cell(target_row, 8).value = status_text
                ws.cell(target_row, 9).value = start_time
                ws.cell(target_row, 10).value = plan_end
                ws.cell(target_row, 11).value = actual_end
                ws.cell(target_row, 12).value = model_str
                ws.cell(target_row, 13).value = route_str
                ws.cell(target_row, 15).value = yes
                ws.cell(target_row, 16).value = yes
                ws.cell(target_row, 17).value = yes

                font10 = Font(size=10)
                for col in range(1, 18):
                    ws.cell(row=target_row, column=col).font = font10

                align_right = Alignment(horizontal='right', vertical='center')
                for col in [2, 9, 10, 11]:
                    ws.cell(row=target_row, column=col).alignment = align_right

                align_left_no_wrap = Alignment(horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=False)
                ws.cell(row=target_row, column=13).alignment = align_left_no_wrap
                ws.column_dimensions['M'].width = 50

                prev_row = target_row - 1
                if prev_row >= 1:
                    src_f_num = ws.cell(prev_row, 6).number_format
                    if src_f_num:
                        ws.cell(target_row, 6).number_format = src_f_num
                    else:
                        ws.cell(target_row, 6).number_format = '0'
                    src_g_num = ws.cell(prev_row, 7).number_format
                    if src_g_num:
                        ws.cell(target_row, 7).number_format = src_g_num
                    else:
                        ws.cell(target_row, 7).number_format = '0'
                else:
                    ws.cell(target_row, 6).number_format = '0'
                    ws.cell(target_row, 7).number_format = '0'

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                st.success("✅ 处理完成！已添加一行新数据。")
                col1, col2, col3 = st.columns(3)
                col1.metric("总架次", total_flights)
                col2.metric("状态", status_text)
                col3.metric("航空器型号数", len(models))

                st.write(f"**开始时间：** {start_time}")
                st.write(f"**计划结束：** {plan_end}")
                st.write(f"**实际结束：** {actual_end if actual_end else '未结束'}")
                st.write(f"**航线：** {route_str}")

                st.download_button(
                    label="⬇️ 下载更新后的模板",
                    data=output,
                    file_name="天成商务航空每日运行跟踪.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"处理失败：{e}")
                st.exception(e)

# ==============================
# 功能 D：通航脚本生成器（完整版，包含所有脚本生成函数）
# ==============================
# ---------- 以下为脚本生成所需的全局常量和辅助函数 ----------
COUNTRIES = [
    "香港", "澳门", "台湾", "蒙古", "朝鲜", "韩国", "日本", "菲律宾", "越南", "老挝",
    "柬埔寨", "缅甸", "泰国", "马来西亚", "文莱", "新加坡", "印度尼西亚", "东帝汶",
    "尼泊尔", "不丹", "孟加拉国", "印度", "巴基斯坦", "斯里兰卡", "马尔代夫",
    "哈萨克斯坦", "吉尔吉斯斯坦", "塔吉克斯坦", "乌兹别克斯坦", "土库曼斯坦",
    "阿富汗", "伊拉克", "伊朗", "叙利亚", "约旦", "黎巴嫩", "以色列", "巴勒斯坦",
    "沙特阿拉伯", "巴林", "卡塔尔", "科威特", "阿联酋", "阿曼", "也门", "格鲁吉亚",
    "亚美尼亚", "阿塞拜疆", "土耳其", "塞浦路斯", "芬兰", "瑞典", "挪威", "冰岛",
    "丹麦", "法罗群岛（丹）", "爱沙尼亚", "拉脱维亚", "立陶宛", "白俄罗斯", "俄罗斯",
    "乌克兰", "摩尔多瓦", "波兰", "捷克", "斯洛伐克", "匈牙利", "德国", "奥地利",
    "瑞士", "列支敦士登", "英国", "爱尔兰", "荷兰", "比利时", "卢森堡", "法国",
    "摩纳哥", "罗马尼亚", "保加利亚", "塞尔维亚", "马其顿", "阿尔巴尼亚", "希腊",
    "斯洛文尼亚", "克罗地亚", "波斯尼亚和墨塞哥维那", "意大利", "梵蒂冈", "圣马力诺",
    "马耳他", "西班牙", "葡萄牙", "安道尔", "埃及", "利比亚", "苏丹", "突尼斯",
    "阿尔及利亚", "摩洛哥", "亚速尔群岛（葡）", "马德拉群岛（葡）", "埃塞俄比亚",
    "厄立特里亚", "索马里", "吉布提", "肯尼亚", "坦桑尼亚", "乌干达", "卢旺达",
    "布隆迪", "塞舌尔", "乍得", "中非", "喀麦隆", "赤道几内亚", "加蓬", "刚果（布）",
    "刚果（金）", "圣多美及普林西比", "毛里塔尼亚", "西撒哈拉", "塞内加尔", "冈比亚",
    "马里", "布基纳法索", "几内亚", "几内亚比绍", "佛得角", "塞拉利昂", "利比里亚",
    "科特迪瓦", "加纳", "多哥", "贝宁", "尼日尔", "加那利群岛（西）", "赞比亚",
    "安哥拉", "津巴布韦", "马拉维", "莫桑比克", "博茨瓦纳", "纳米比亚", "南非",
    "斯威士兰", "莱索托", "马达加斯加", "科摩罗", "毛里求斯", "留尼旺（法）",
    "圣赫勒拿（英）", "澳大利亚", "新西兰", "巴布亚新几内亚", "所罗门群岛",
    "瓦努阿图", "密克罗尼西亚", "马绍尔群岛", "帕劳", "瑙鲁", "基里巴斯", "图瓦卢",
    "萨摩亚", "斐济群岛", "汤加", "库克群岛（新）", "关岛（美）", "新喀里多尼亚（法）",
    "法属波利尼西亚", "皮特凯恩岛（英）", "瓦利斯与富图纳（法）", "纽埃（新）",
    "托克劳（新）", "美属萨摩亚", "北马里亚纳（美）", "加拿大", "美国", "墨西哥",
    "格陵兰（丹）", "危地马拉", "伯利兹", "萨尔瓦多", "洪都拉斯", "尼加拉瓜",
    "哥斯达黎加", "巴拿马", "巴哈马", "古巴", "牙买加", "海地", "多米尼加共和国",
    "安提瓜和巴布达", "圣基茨和尼维斯", "多米尼克", "圣卢西亚", "圣文森特和格林纳丁斯",
    "格林纳达", "巴巴多斯", "特立尼达和多巴哥", "波多黎各（美）", "英属维尔京群岛",
    "美属维尔京群岛", "安圭拉（英）", "蒙特塞拉特（英）", "瓜德罗普（法）",
    "马提尼克（法）", "荷属安的列斯", "阿鲁巴（荷）", "特克斯和凯科斯群岛（英）",
    "开曼群岛（英）", "百慕大（英）", "哥伦比亚", "委内瑞拉", "圭亚那", "法属圭亚那",
    "苏里南", "厄瓜多尔", "秘鲁", "玻利维亚", "智利", "阿根廷", "乌拉圭", "巴拉圭",
    "其他"
]

ABBR_TO_COUNTRY = {
    "印尼": "印度尼西亚",
    "台北": "台湾",
    "高雄": "台湾",
    "台中": "台湾",
    "花莲": "台湾",
    "台东": "台湾",
    "嘉义": "台湾",
    "台南": "台湾",
}

TAIWAN_CITIES = [
    "台北松山", "台北桃园", "高雄小港", "台中清泉岗", "花莲", "台东", "嘉义", "台南", "马公", "金门", "马祖"
]

CITY_TO_PROVINCE = {
    "北京": "北京", "上海": "上海", "天津": "天津", "重庆": "重庆",
    "广州": "广东", "深圳": "广东", "珠海": "广东", "汕头": "广东", "佛山": "广东", "江门": "广东",
    "湛江": "广东", "茂名": "广东", "肇庆": "广东", "惠州": "广东", "梅州": "广东", "汕尾": "广东",
    "河源": "广东", "阳江": "广东", "清远": "广东", "东莞": "广东", "中山": "广东", "潮州": "广东",
    "揭阳": "广东", "云浮": "广东",
    "南京": "江苏", "无锡": "江苏", "徐州": "江苏", "常州": "江苏", "苏州": "江苏", "南通": "江苏",
    "连云港": "江苏", "淮安": "江苏", "盐城": "江苏", "扬州": "江苏", "镇江": "江苏", "泰州": "江苏",
    "宿迁": "江苏",
    "杭州": "浙江", "宁波": "浙江", "温州": "浙江", "嘉兴": "浙江", "湖州": "浙江", "绍兴": "浙江",
    "金华": "浙江", "衢州": "浙江", "舟山": "浙江", "台州": "浙江", "丽水": "浙江",
    "合肥": "安徽", "芜湖": "安徽", "蚌埠": "安徽", "淮南": "安徽", "马鞍山": "安徽", "淮北": "安徽",
    "铜陵": "安徽", "安庆": "安徽", "黄山": "安徽", "滁州": "安徽", "阜阳": "安徽", "宿州": "安徽",
    "六安": "安徽", "亳州": "安徽", "池州": "安徽", "宣城": "安徽",
    "福州": "福建", "厦门": "福建", "莆田": "福建", "三明": "福建", "泉州": "福建", "漳州": "福建",
    "南平": "福建", "龙岩": "福建", "宁德": "福建",
    "南昌": "江西", "景德镇": "江西", "萍乡": "江西", "九江": "江西", "新余": "江西", "鹰潭": "江西",
    "赣州": "江西", "吉安": "江西", "宜春": "江西", "抚州": "江西", "上饶": "江西",
    "济南": "山东", "青岛": "山东", "淄博": "山东", "枣庄": "山东", "东营": "山东", "烟台": "山东",
    "潍坊": "山东", "济宁": "山东", "泰安": "山东", "威海": "山东", "日照": "山东", "临沂": "山东",
    "德州": "山东", "聊城": "山东", "滨州": "山东", "菏泽": "山东",
    "郑州": "河南", "开封": "河南", "洛阳": "河南", "平顶山": "河南", "安阳": "河南", "鹤壁": "河南",
    "新乡": "河南", "焦作": "河南", "濮阳": "河南", "许昌": "河南", "漯河": "河南", "三门峡": "河南",
    "南阳": "河南", "商丘": "河南", "信阳": "河南", "周口": "河南", "驻马店": "河南",
    "武汉": "湖北", "黄石": "湖北", "十堰": "湖北", "宜昌": "湖北", "襄阳": "湖北", "鄂州": "湖北",
    "荆门": "湖北", "孝感": "湖北", "荆州": "湖北", "黄冈": "湖北", "咸宁": "湖北", "随州": "湖北",
    "长沙": "湖南", "株洲": "湖南", "湘潭": "湖南", "衡阳": "湖南", "邵阳": "湖南", "岳阳": "湖南",
    "常德": "湖南", "张家界": "湖南", "益阳": "湖南", "郴州": "湖南", "永州": "湖南", "怀化": "湖南",
    "娄底": "湖南",
    "成都": "四川", "自贡": "四川", "攀枝花": "四川", "泸州": "四川", "德阳": "四川", "绵阳": "四川",
    "广元": "四川", "遂宁": "四川", "内江": "四川", "乐山": "四川", "南充": "四川", "眉山": "四川",
    "宜宾": "四川", "广安": "四川", "达州": "四川", "雅安": "四川", "巴中": "四川", "资阳": "四川",
    "贵阳": "贵州", "六盘水": "贵州", "遵义": "贵州", "安顺": "贵州", "毕节": "贵州", "铜仁": "贵州",
    "昆明": "云南", "曲靖": "云南", "玉溪": "云南", "保山": "云南", "昭通": "云南", "丽江": "云南",
    "普洱": "云南", "临沧": "云南",
    "西安": "陕西", "铜川": "陕西", "宝鸡": "陕西", "咸阳": "陕西", "渭南": "陕西", "延安": "陕西",
    "汉中": "陕西", "榆林": "陕西", "安康": "陕西", "商洛": "陕西",
    "兰州": "甘肃", "嘉峪关": "甘肃", "金昌": "甘肃", "白银": "甘肃", "天水": "甘肃", "武威": "甘肃",
    "张掖": "甘肃", "平凉": "甘肃", "酒泉": "甘肃", "庆阳": "甘肃", "定西": "甘肃", "陇南": "甘肃",
    "西宁": "青海", "海东": "青海",
    "银川": "宁夏", "石嘴山": "宁夏", "吴忠": "宁夏", "固原": "宁夏", "中卫": "宁夏",
    "乌鲁木齐": "新疆", "克拉玛依": "新疆", "吐鲁番": "新疆", "哈密": "新疆",
    "伊犁": "新疆",
    "伊犁伊宁": "新疆",
    "拉萨": "西藏", "日喀则": "西藏", "昌都": "西藏", "林芝": "西藏", "山南": "西藏", "那曲": "西藏",
    "呼和浩特": "内蒙古", "包头": "内蒙古", "乌海": "内蒙古", "赤峰": "内蒙古", "通辽": "内蒙古",
    "鄂尔多斯": "内蒙古", "呼伦贝尔": "内蒙古", "巴彦淖尔": "内蒙古", "乌兰察布": "内蒙古",
    "南宁": "广西", "柳州": "广西", "桂林": "广西", "梧州": "广西", "北海": "广西", "防城港": "广西",
    "钦州": "广西", "贵港": "广西", "玉林": "广西", "百色": "广西", "贺州": "广西", "河池": "广西",
    "来宾": "广西", "崇左": "广西",
    "海口": "海南", "三亚": "海南", "三沙": "海南", "儋州": "海南",
    "石家庄": "河北", "唐山": "河北", "秦皇岛": "河北", "邯郸": "河北", "邢台": "河北", "保定": "河北",
    "张家口": "河北", "承德": "河北", "沧州": "河北", "廊坊": "河北", "衡水": "河北",
    "太原": "山西", "大同": "山西", "阳泉": "山西", "长治": "山西", "晋城": "山西", "朔州": "山西",
    "晋中": "山西", "运城": "山西", "忻州": "山西", "临汾": "山西", "吕梁": "山西",
    "沈阳": "辽宁", "大连": "辽宁", "鞍山": "辽宁", "抚顺": "辽宁", "本溪": "辽宁", "丹东": "辽宁",
    "锦州": "辽宁", "营口": "辽宁", "阜新": "辽宁", "辽阳": "辽宁", "盘锦": "辽宁", "铁岭": "辽宁",
    "朝阳": "辽宁", "葫芦岛": "辽宁",
    "长春": "吉林", "吉林": "吉林", "四平": "吉林", "辽源": "吉林", "通化": "吉林", "白山": "吉林",
    "松原": "吉林", "白城": "吉林",
    "哈尔滨": "黑龙江", "齐齐哈尔": "黑龙江", "鸡西": "黑龙江", "鹤岗": "黑龙江", "双鸭山": "黑龙江",
    "大庆": "黑龙江", "伊春": "黑龙江", "佳木斯": "黑龙江", "七台河": "黑龙江", "牡丹江": "黑龙江",
    "黑河": "黑龙江", "绥化": "黑龙江",
}

DOMESTIC_KEYWORDS = list(CITY_TO_PROVINCE.keys())

DEFAULT_DETAIL_MAP = {
    "北京首都": {"province": "北京", "district": "顺义区"},
    "北京大兴": {"province": "北京", "district": "大兴区"},
    "天津滨海": {"province": "天津", "district": "滨海新区"},
    "上海虹桥": {"province": "上海", "district": "闵行区"},
    "上海浦东": {"province": "上海", "district": "浦东新区"},
    "重庆江北": {"province": "重庆", "district": "江北区"},
    "乌鲁木齐天山": {"province": "新疆", "district": "乌鲁木齐"},
    "伊犁伊宁": {"province": "新疆", "district": "伊犁"},
}

def parse_flight_time(time_str):
    try:
        parts = time_str.split(':')
        hours = int(parts[0])
        minutes = int(parts[1])
        return hours, minutes
    except:
        return 0, 0

def extract_country(city_name):
    if city_name.startswith("印尼"):
        return "印度尼西亚"
    parts = re.split(r'[\s\-]', city_name)
    if parts:
        first_part = parts[0]
        if first_part in ABBR_TO_COUNTRY:
            return ABBR_TO_COUNTRY[first_part]
    for tw_city in TAIWAN_CITIES:
        if tw_city in city_name:
            return "台湾"
    for country in COUNTRIES:
        if country in city_name:
            return country
    if parts:
        return parts[0]
    return city_name

def get_province_from_city(city):
    for keyword, province in CITY_TO_PROVINCE.items():
        if keyword in city:
            return province
    return city.split()[0] if city.split() else city

def extract_district_from_city(city):
    for keyword in CITY_TO_PROVINCE.keys():
        if keyword in city:
            return keyword
    district = city.split()[0] if city.split() else city
    if district.endswith('机场'):
        district = district[:-2]
    return district

def build_city_mappings(df, custom_detail_map):
    cities = set()
    for _, row in df.iterrows():
        dep = str(row["出发城市"]).strip()
        arr = str(row["到达城市"]).strip()
        cities.add(dep)
        cities.add(arr)

    detail_map = {**DEFAULT_DETAIL_MAP, **custom_detail_map}
    city_map = {}

    for city in cities:
        if any(tw in city for tw in TAIWAN_CITIES):
            city_map[city] = "台湾"
            continue

        if city in detail_map:
            province = detail_map[city]["province"]
            city_map[city] = province
            continue

        is_domestic = any(kw in city for kw in DOMESTIC_KEYWORDS)
        if is_domestic:
            province = get_province_from_city(city)
            district = extract_district_from_city(city)
            detail_map[city] = {"province": province, "district": district}
            city_map[city] = province
        else:
            country = extract_country(city)
            city_map[city] = country

    return city_map, detail_map

# ---------- 以下三个函数为脚本生成核心，完整包含 ----------
def generate_base_script(city_map_json, detail_map_json, domestic_keywords_json):
    return f"""
// ================= 公共辅助函数（基础脚本） =================
// 生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
// =========================================================

function sleep(ms) {{ return new Promise(r => setTimeout(r, ms)); }}

async function getMainDoc() {{
    let iframe = document.querySelector('#main');
    if (iframe) {{
        let doc = iframe.contentDocument;
        if (doc && doc.querySelector('body')) {{
            return doc;
        }}
    }}
    console.log('未找到 iframe #main，使用顶层文档');
    return document;
}}

async function waitForElement(selector, timeout = 15000, isXPath = true) {{
    const start = Date.now();
    while (Date.now() - start < timeout) {{
        const doc = await getMainDoc();
        if (!doc) return null;
        let el;
        if (isXPath) {{
            el = doc.evaluate(selector, doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        }} else {{
            el = doc.querySelector(selector);
        }}
        if (el) return el;
        await sleep(300);
    }}
    console.warn(`⚠️ 等待元素超时: ${{selector}}`);
    return null;
}}

function setDateInput(inputEl, dateStr) {{
    if (!inputEl) return false;
    inputEl.value = dateStr;
    inputEl.dispatchEvent(new Event('input', {{ bubbles: true }}));
    inputEl.dispatchEvent(new Event('change', {{ bubbles: true }}));
    inputEl.blur();
    return true;
}}

async function setSelectValue(selectEl, valueText) {{
    if (!selectEl) return false;
    for (let i = 0; i < selectEl.options.length; i++) {{
        const opt = selectEl.options[i];
        if (opt.text === valueText || opt.text.includes(valueText)) {{
            selectEl.selectedIndex = i;
            selectEl.dispatchEvent(new Event('change', {{ bubbles: true }}));
            await sleep(300);
            console.log(`✅ 已选择 "${{valueText}}"`);
            return true;
        }}
    }}
    console.warn(`⚠️ 未找到选项 "${{valueText}}"`);
    return false;
}}

function getAirportNameFromCity(city) {{
    const firstPart = city.split(/\\s+/)[0];
    return firstPart + "机场";
}}

function setNumberInput(inputEl, value) {{
    if (!inputEl) return false;
    inputEl.value = value;
    inputEl.dispatchEvent(new Event('input', {{ bubbles: true }}));
    inputEl.dispatchEvent(new Event('change', {{ bubbles: true }}));
    inputEl.blur();
    return true;
}}

const CITY_MAP_RAW = {city_map_json};
const CITY_MAP_CORRECTED = {{
    ...CITY_MAP_RAW,
    "印尼雅加达哈林": "印度尼西亚",
    "印尼巴厘岛": "印度尼西亚",
    "台北松山": "台湾",
    "台北桃园": "台湾",
    "高雄小港": "台湾",
    "台中清泉岗": "台湾",
    "花莲": "台湾",
    "台东": "台湾",
    "嘉义": "台湾",
    "台南": "台湾",
    "马公": "台湾",
    "金门": "台湾",
    "马祖": "台湾"
}};
const CITY_MAP = CITY_MAP_CORRECTED;
const DOMESTIC_KEYWORDS = {domestic_keywords_json};

function getLocationInfo(city) {{
    const isDomestic = DOMESTIC_KEYWORDS.some(keyword => city.includes(keyword));
    if (isDomestic) {{
        let region = CITY_MAP[city];
        if (!region) {{ const match = city.match(/^([^\\s\\-]+)/); region = match ? match[1] : city; }}
        return {{ zone: "境内", region: region }};
    }} else {{
        let country = CITY_MAP[city];
        if (!country) {{
            const parts = city.split(/[\\s\\-]/);
            if (parts.length > 0) {{
                const first = parts[0];
                if (first === "印尼") country = "印度尼西亚";
                else if (["台北","高雄","台中","花莲","台东","嘉义","台南"].includes(first)) country = "台湾";
                else country = first;
            }} else {{
                country = city;
            }}
        }}
        return {{ zone: "境外", region: country }};
    }}
}}

const CITY_DETAIL_MAP = {detail_map_json};

async function fillSegmentSelects(container, city) {{
    const selects = container.querySelectorAll('select');
    if (selects.length < 2) {{
        console.warn('航段容器内 select 数量不足');
        return false;
    }}
    
    const info = getLocationInfo(city);
    if (info.zone === "境外") {{
        await setSelectValue(selects[0], info.zone);
        await setSelectValue(selects[1], info.region);
        return true;
    }}
    
    const detail = CITY_DETAIL_MAP[city];
    if (!detail) {{
        console.warn(`未找到城市 ${{city}} 的详细映射，将使用降级处理`);
        await setSelectValue(selects[0], "境内");
        await setSelectValue(selects[1], info.region);
        if (selects.length >= 3) {{
            const thirdSelect = selects[2];
            let chooseBtn = null;
            const possibleButtons = container.querySelectorAll('button, div, span');
            for (let el of possibleButtons) {{
                if (el.innerText && el.innerText.includes('请选择')) {{
                    chooseBtn = el;
                    break;
                }}
            }}
            if (chooseBtn) {{
                chooseBtn.click();
                await sleep(1000);
            }}
            await sleep(1000);
            if (thirdSelect.options.length > 1) {{
                console.warn(`未找到区县选项，第三个下拉框将保持当前选择（默认为第一个选项）`);
            }} else {{
                console.warn('第三个下拉框选项不足');
            }}
        }}
        return true;
    }}
    
    await setSelectValue(selects[0], "境内");
    await setSelectValue(selects[1], detail.province);
    console.log(`等待第三个下拉框选项加载 (${{detail.district}})...`);
    await sleep(1500);
    
    const newSelects = container.querySelectorAll('select');
    if (newSelects.length < 3) {{
        console.warn('重新获取后第三个下拉框不存在');
        return false;
    }}
    const thirdSelect = newSelects[2];
    console.log('第三个下拉框当前选项:', Array.from(thirdSelect.options).map(o => o.text));
    
    let targetIndex = -1;
    for (let i = 0; i < thirdSelect.options.length; i++) {{
        if (thirdSelect.options[i].text.includes(detail.district)) {{
            targetIndex = i;
            break;
        }}
    }}
    
    if (targetIndex !== -1) {{
        thirdSelect.selectedIndex = targetIndex;
        thirdSelect.dispatchEvent(new Event('change', {{ bubbles: true }}));
        console.log(`已选择第三个下拉框: ${{thirdSelect.options[targetIndex].text}}`);
    }} else {{
        console.warn(`未找到区县选项: ${{detail.district}}，请手动选择或补充映射。`);
    }}
    await sleep(500);
    return true;
}}

async function fillFirstSegmentSelects(city) {{
    const container = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[1]/div[1]/div/div', 5000, true);
    if (!container) {{
        console.warn('未找到第一个航段容器');
        return false;
    }}
    return fillSegmentSelects(container, city);
}}

async function fillFirstSegmentTime(record) {{
    const hourInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[1]/div[2]/div/input[1]', 5000, true);
    const minuteInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[1]/div[2]/div/input[2]', 5000, true);
    const flightsInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[1]/div[3]/div/input', 5000, true);
    if (hourInput && minuteInput && flightsInput) {{
        hourInput.value = record.flight_hours.toString().padStart(2, '0');
        minuteInput.value = record.flight_minutes.toString().padStart(2, '0');
        flightsInput.value = 1;
        hourInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        minuteInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        flightsInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        console.log(`已填入第一个航段时间: ${{record.flight_hours}}:${{record.flight_minutes}}, 架次: 1`);
        return true;
    }}
    console.warn('无法填入第一个航段时间和架次');
    return false;
}}

async function fillSecondSegmentSelects(city) {{
    const container = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[2]/div[1]/div/div', 5000, true);
    if (!container) {{
        console.warn('未找到第二个航段容器');
        return false;
    }}
    return fillSegmentSelects(container, city);
}}

async function fillSecondSegmentTime() {{
    const hourInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[2]/div[2]/div/input[1]', 5000, true);
    const minuteInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[2]/div[2]/div/input[2]', 5000, true);
    const flightsInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[2]/div[3]/div/input', 5000, true);
    if (hourInput && minuteInput && flightsInput) {{
        hourInput.value = '00';
        minuteInput.value = '00';
        flightsInput.value = 0;
        hourInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        minuteInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        flightsInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        console.log('已填入第二个航段时间: 00:00, 架次: 0');
        return true;
    }}
    console.warn('无法填入第二个航段时间和架次');
    return false;
}}

function formatRegNumber(reg) {{
    if (reg.includes('-')) return reg;
    const match = reg.match(/^([A-Z]+)(\\d+[A-Z]*)$/);
    if (match) return `${{match[1]}}-${{match[2]}}`;
    return reg;
}}

async function selectAircraft(reg) {{
    const regForSelect = formatRegNumber(reg);
    console.log(`尝试选择飞机: ${{regForSelect}}`);
    
    const airbox = await waitForElement('//*[@id="airbox"]', 10000, true);
    if (!airbox) {{
        console.warn('未找到飞机选择对话框');
        return false;
    }}
    const doc = await getMainDoc();
    
    let span = doc.evaluate(`//span[text()='${{regForSelect}}']`, doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
    if (!span) {{
        span = doc.evaluate(`//span[contains(text(), '${{regForSelect}}')]`, doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
    }}
    if (!span) {{
        console.warn(`未找到包含注册号 ${{regForSelect}} 的 span 元素`);
        return false;
    }}
    const li = span.closest('li');
    if (!li) {{
        console.warn(`未找到注册号 ${{regForSelect}} 对应的 li`);
        return false;
    }}
    const checkbox = li.querySelector('input[type="checkbox"]');
    if (!checkbox) {{
        console.warn(`未找到注册号 ${{regForSelect}} 的复选框`);
        return false;
    }}
    checkbox.click();
    console.log('已勾选飞机复选框');
    await sleep(300);
    
    let closeBtn = doc.evaluate('//*[@id="close7"]', doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
    if (!closeBtn) {{
        closeBtn = doc.evaluate('//button[contains(text(), "关闭")]', doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
    }}
    if (!closeBtn) {{
        closeBtn = doc.evaluate('//button[contains(text(), "确定")]', doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
    }}
    if (closeBtn) {{
        closeBtn.click();
        console.log('已关闭选择对话框');
        await sleep(500);
        return true;
    }} else {{
        console.warn('未找到关闭按钮，尝试按 ESC 关闭');
        document.dispatchEvent(new KeyboardEvent('keydown', {{ key: 'Escape' }}));
        await sleep(500);
        return false;
    }}
}}

async function waitForDialogConfirmButton(timeout = 15000) {{
    const start = Date.now();
    while (Date.now() - start < timeout) {{
        let btn = document.evaluate('//a[contains(text(), "确定")]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        if (!btn) btn = document.evaluate('//button[contains(text(), "确定")]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        if (btn) return btn;
        try {{
            const doc = await getMainDoc();
            btn = doc.evaluate('//a[contains(text(), "确定")]', doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
            if (!btn) btn = doc.evaluate('//button[contains(text(), "确定")]', doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
            if (btn) return btn;
        }} catch(e) {{}}
        await sleep(300);
    }}
    return null;
}}

async function handleAirportBlock(blockIndex, city, label) {{
    let firstSelectXPath, secondSelectXPath;
    if (blockIndex === 1) {{
        firstSelectXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[1]/div[2]/div/div[1]/div/select[1]';
        secondSelectXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[1]/div[2]/div/div[1]/div/select[2]';
    }} else {{
        firstSelectXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[2]/div[2]/div/div[1]/div/select[1]';
        secondSelectXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[2]/div[2]/div/div[1]/div/select[2]';
    }}
    console.log(`⏳ 处理 ${{label}} 区块...`);
    const firstSelect = await waitForElement(firstSelectXPath, 10000);
    if (!firstSelect) {{ console.error(`❌ 未找到 ${{label}} 第一个选择框`); return false; }}
    const targetValue = blockIndex === 1 ? '起飞机场' : '降落机场';
    await setSelectValue(firstSelect, targetValue);
    
    const secondSelect = await waitForElement(secondSelectXPath, 10000);
    if (!secondSelect) {{ console.error(`❌ 未找到 ${{label}} 第二个选择框`); return false; }}
    
    const info = getLocationInfo(city);
    if (info.zone === "境内") {{
        console.log(`🛬 ${{label}} 境内机场: ${{city}}，尝试选择匹配的机场...`);
        let selected = await setSelectValue(secondSelect, city);
        if (!selected) {{
            const detail = CITY_DETAIL_MAP[city];
            if (detail && detail.district) {{
                console.log(`尝试使用区县名 "${{detail.district}}" 进行匹配...`);
                selected = await setSelectValue(secondSelect, detail.district);
            }}
        }}
        if (!selected) {{
            console.error(`❌ 未找到匹配的机场选项 (尝试了 "${{city}}" 和区县名)`);
            return false;
        }}
        return true;
    }} else {{
        console.log(`🛫 ${{label}} 境外机场: ${{city}}，选择“其它”...`);
        const otherSelected = await setSelectValue(secondSelect, '其它');
        if (!otherSelected) {{ console.error(`❌ 无法选择“其它”`); return false; }}
        await sleep(800);
        let zoneSelectXPath, zoneSelect2XPath, airportNameInputXPath;
        if (blockIndex === 2) {{
            zoneSelectXPath = '/html/body/div/div[1]/div[3]/div/div[2]/form/div[11]/div[2]/div[2]/div/div[2]/div/select[1]';
            zoneSelect2XPath = '/html/body/div/div[1]/div[3]/div/div[2]/form/div[11]/div[2]/div[2]/div/div[2]/div/select[2]';
            airportNameInputXPath = '/html/body/div/div[1]/div[3]/div/div[2]/form/div[11]/div[2]/div[2]/div/div[2]/div/input';
        }} else {{
            zoneSelectXPath = `/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[1]/div[2]/div/div[2]/div/select[1]`;
            zoneSelect2XPath = `/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[1]/div[2]/div/div[2]/div/select[2]`;
            airportNameInputXPath = `/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[1]/div[2]/div/div[2]/div/input`;
        }}
        const zoneSelect = await waitForElement(zoneSelectXPath, 10000);
        if (zoneSelect) await setSelectValue(zoneSelect, '境外');
        const zoneSelect2 = await waitForElement(zoneSelect2XPath, 10000);
        if (zoneSelect2) await setSelectValue(zoneSelect2, '境外');
        const airportNameInput = await waitForElement(airportNameInputXPath, 10000);
        if (airportNameInput) {{
            const airportName = getAirportNameFromCity(city);
            console.log(`📝 填入 ${{label}} 机场名称: ${{airportName}}`);
            setNumberInput(airportNameInput, airportName);
        }}
        return true;
    }}
}}

async function handleSecondFlightTime() {{
    const hourXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[2]/div[2]/div/div[5]/div/input[1]';
    const minuteXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[2]/div[2]/div/div[5]/div/input[2]';
    const countXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[2]/div[2]/div/div[6]/div/input';
    const hourInput = await waitForElement(hourXPath, 10000);
    const minuteInput = await waitForElement(minuteXPath, 10000);
    const countInput = await waitForElement(countXPath, 10000);
    if (hourInput && minuteInput && countInput) {{
        console.log('⏱️ 填入第二个飞行时间: 00:00，飞行架次数: 0');
        setNumberInput(hourInput, '0');
        setNumberInput(minuteInput, '0');
        setNumberInput(countInput, '0');
        return true;
    }} else {{
        console.warn('⚠️ 未找到第二个飞行时间输入框');
        return false;
    }}
}}

async function handleDetailArea(depCity, arrCity) {{
    const detailXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[16]/div/input';
    const detailInput = await waitForElement(detailXPath, 10000);
    if (!detailInput) {{ console.error('❌ 未找到详细作业区输入框'); return false; }}
    const depAirport = depCity + "机场";
    const arrAirport = arrCity + "机场";
    const detailText = `${{depAirport}}-${{arrAirport}}`;
    console.log(`📝 填入详细作业区: ${{detailText}}`);
    setNumberInput(detailInput, detailText);
    return true;
}}

async function waitForReturnToList(timeout = 300000) {{
    const start = Date.now();
    console.log('⏳ 等待您手动点击“提交”后返回列表页...');
    while (Date.now() - start < timeout) {{
        const doc = await getMainDoc();
        if (!doc) return false;
        const rows = doc.querySelectorAll(ROW_SELECTOR);
        if (rows.length > 0) {{
            console.log('✅ 已返回列表页');
            return true;
        }}
        await sleep(2000);
    }}
    console.error('❌ 等待超时，未返回列表页');
    return false;
}}

async function ensureListPage() {{
    const btn = await waitForElement('input.query.yuanjiao', 15000, false);
    if (btn) return true;
    console.log('当前不在列表页，尝试关闭可能遗留的对话框...');
    const doc = await getMainDoc();
    if (!doc) return false;
    let closeBtn = doc.evaluate('//button[contains(text(), "取消")]', doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
    if (!closeBtn) closeBtn = doc.evaluate('//button[contains(text(), "关闭")]', doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
    if (!closeBtn) closeBtn = doc.evaluate('//button[contains(text(), "返回")]', doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
    if (closeBtn) {{
        closeBtn.click();
        console.log('已点击关闭按钮，等待返回列表页');
        await sleep(1000);
        const backBtn = await waitForElement('input.query.yuanjiao', 10000, false);
        return backBtn !== null;
    }} else {{
        console.warn('未找到返回按钮，请手动关闭对话框后继续（脚本将等待5秒）');
        await sleep(5000);
        const backBtn = await waitForElement('input.query.yuanjiao', 5000, false);
        return backBtn !== null;
    }}
}}
"""

def generate_daily_script(records, city_map_json, detail_map_json, domestic_keywords_json):
    js_data = json.dumps(records, ensure_ascii=False, indent=4)
    script = f"""
// ================= 自动生成的飞行计划脚本（当日计划专用） =================
// 当日待处理计划数: {len(records)}
// =========================================================

const IFRAME_ID = 'main';
const ROW_SELECTOR = 'table tbody:nth-of-type(2) tr';
const REG_SELECTOR = 'td:nth-child(6) div';
const SEGMENT_SELECTOR = 'td:nth-child(7) div';
const DATE_SELECTOR = 'td:nth-child(9)';

const excelData = {js_data};

function normalizeReg(reg) {{
    return reg.replace(/[-\s]/g, '').trim();
}}

function getPlanKey(plan) {{
    return `${{plan["飞机注册号"]}}_${{plan["出发日期"]}}_${{plan["出发城市"]}}_${{plan["到达城市"]}}`;
}}

async function waitForTable() {{
    const start = Date.now();
    while (Date.now() - start < 10000) {{
        const doc = await getMainDoc();
        if (!doc) return null;
        const rows = doc.querySelectorAll(ROW_SELECTOR);
        if (rows.length > 0) return rows;
        await sleep(300);
    }}
    return null;
}}

function extractCityKeywords(segText) {{
    let parts = segText.split(',');
    if (parts.length < 2) return null;
    let depPart = parts[0].trim();
    let arrPart = parts[1].trim();
    depPart = depPart.replace(/^(境内|境外)-/, '');
    arrPart = arrPart.replace(/^(境内|境外)-/, '');
    let extract = (part) => {{
        let segments = part.split('-');
        let keywords = [];
        for (let seg of segments) {{
            let words = seg.split(/\\s+/);
            for (let w of words) if (w) keywords.push(w);
        }}
        return keywords;
    }};
    return {{ depKeywords: extract(depPart), arrKeywords: extract(arrPart) }};
}}

async function tryMatchPlan(plan) {{
    const rows = await waitForTable();
    if (!rows) return null;
    const regTarget = normalizeReg(plan["飞机注册号"]);
    const dateTarget = plan["出发日期"];
    const depCity = plan["出发城市"];
    const arrCity = plan["到达城市"];
    const depInfo = getLocationInfo(depCity);
    const arrInfo = getLocationInfo(arrCity);
    const depKey = depInfo.zone === "境外" ? depInfo.region : depCity;
    const arrKey = arrInfo.zone === "境外" ? arrInfo.region : arrCity;

    for (let i = 0; i < rows.length; i++) {{
        const row = rows[i];
        const regEl = row.querySelector(REG_SELECTOR);
        const regNo = regEl ? normalizeReg(regEl.innerText.trim()) : null;
        if (regNo !== regTarget) continue;
        const dateCell = row.querySelector(DATE_SELECTOR);
        const webDate = dateCell ? dateCell.innerText.trim() : '';
        if (webDate !== dateTarget) continue;
        const segEl = row.querySelector(SEGMENT_SELECTOR);
        if (!segEl) continue;
        const segText = segEl.innerText.trim();
        const kw = extractCityKeywords(segText);
        if (!kw) continue;
        const depMatch = kw.depKeywords.some(kw => depKey.includes(kw));
        const arrMatch = kw.arrKeywords.some(kw => arrKey.includes(kw));
        if (depMatch && arrMatch) {{
            return row;
        }}
    }}
    return null;
}}

// ---------- 修改点1：waitForManualExecute 支持空格/回车跳过 ----------
async function waitForManualExecute(plan) {{
    console.log(`%c⏳ 等待手动操作：请点击计划【${{plan["飞机注册号"]}} - ${{plan["出发城市"]}} → ${{plan["到达城市"]}}】对应的“执行”按钮`, 'background: #ff9800; color: white; font-size: 14px; padding: 4px 8px; border-radius: 4px');
    console.log(`%c💡 按 空格键 或 回车键 可跳过此计划，继续处理下一个`, 'background: #2196F3; color: white; font-size: 12px; padding: 2px 6px; border-radius: 4px');
    console.log(`计划详情: 日期=${{plan["出发日期"]}}, 用途=${{plan["用途"]}}, 飞行时间=${{plan["实际飞行时间"]}}`);
    
    let skipPressed = false;
    const keyHandler = (e) => {{
        if (e.key === ' ' || e.key === 'Enter') {{
            e.preventDefault();
            skipPressed = true;
            document.removeEventListener('keydown', keyHandler);
            console.log(`⏭️ 检测到空格/回车，跳过计划 ${{plan["飞机注册号"]}}`);
        }}
    }};
    document.addEventListener('keydown', keyHandler);

    const startDateXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[9]/div/input';
    let startInput = null;
    let attempts = 0;
    while (attempts < 180 && !skipPressed) {{
        await sleep(500);
        if (skipPressed) break;
        const doc = await getMainDoc();
        if (!doc) continue;
        startInput = doc.evaluate(startDateXPath, doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        if (startInput) break;
        attempts++;
    }}
    document.removeEventListener('keydown', keyHandler);

    if (skipPressed) {{
        console.log(`⏭️ 已跳过计划：${{plan["飞机注册号"]}}`);
        return null;
    }}
    if (!startInput) {{
        console.error('❌ 等待超时，未检测到表单加载，将跳过此计划');
        return null;
    }}
    console.log('✅ 检测到表单已加载，继续填写');
    return await getMainDoc();
}}

// ---------- 修改点2：processExistingPlanWithManual 处理跳过返回值 ----------
async function processExistingPlanWithManual(row, plan) {{
    console.log(`\\n🔧 开始处理计划（手动执行后自动填写）：机号 ${{plan["飞机注册号"]}}`);
    
    const doc = await waitForManualExecute(plan);
    if (!doc) {{
        console.log(`⏭️ 计划已被跳过或超时，继续下一个。`);
        return false;
    }}

    // 等待日期输入框（再次确认）
    const startDateXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[9]/div/input';
    const endDateXPath   = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[10]/div/input';
    console.log('⏳ 等待日期输入框...');
    const startInput = await waitForElement(startDateXPath, 15000);
    const endInput   = await waitForElement(endDateXPath, 15000);
    if (!startInput || !endInput) {{
        console.error('❌ 未找到日期输入框，可能表单未正确加载');
        return false;
    }}
    const startDate = plan["出发日期"];
    const endDate   = plan["到达日期"];
    console.log(`📅 填入作业开始日期: ${{startDate}}`);
    console.log(`📅 填入作业结束日期: ${{endDate}}`);
    setDateInput(startInput, startDate);
    setDateInput(endInput, endDate);

    const depCity = plan["出发城市"];
    const arrCity = plan["到达城市"];
    if (!(await handleAirportBlock(1, depCity, "起飞"))) return false;
    if (!(await handleAirportBlock(2, arrCity, "降落"))) return false;

    const actualFlightTime = plan["实际飞行时间"];
    if (actualFlightTime && actualFlightTime.includes(':')) {{
        const flightHourXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[1]/div[2]/div/div[5]/div/input[1]';
        const flightMinuteXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[1]/div[2]/div/div[5]/div/input[2]';
        const flightCountXPath = '/html/body/div[1]/div/div[3]/div/div[2]/form/div[11]/div[1]/div[2]/div/div[6]/div/input';
        const hourInput = await waitForElement(flightHourXPath, 10000);
        const minuteInput = await waitForElement(flightMinuteXPath, 10000);
        const countInput = await waitForElement(flightCountXPath, 10000);
        if (hourInput && minuteInput && countInput) {{
            const [hour, minute] = actualFlightTime.split(':');
            console.log(`⏱️ 填入第一个飞行时间: ${{hour}}:${{minute}}，飞行架次数: 1`);
            setNumberInput(hourInput, hour);
            setNumberInput(minuteInput, minute);
            setNumberInput(countInput, '1');
        }} else console.warn('⚠️ 未找到第一个飞行时间输入框');
    }} else console.warn(`⚠️ 实际飞行时间 "${{actualFlightTime}}" 格式不正确`);

    await handleSecondFlightTime();
    await handleDetailArea(depCity, arrCity);

    console.log('✅ 所有字段填写完成，请手动点击“提交”按钮。');
    const returned = await waitForReturnToList();
    if (!returned) {{
        console.error('❌ 未能检测到返回列表页，请手动返回后继续运行脚本。');
        return false;
    }}
    return true;
}}

// ---------- runDailyPlans 保持不变 ----------
async function runDailyPlans() {{
    console.log('🚀 开始执行当日计划自动化流程...');
    let processedCount = 0;
    const processedKeys = new Set();

    for (let plan of excelData) {{
        const key = getPlanKey(plan);
        if (processedKeys.has(key)) continue;
        processedKeys.add(key);
        processedCount++;
        console.log(`\\n========== 处理第 ${{processedCount}} 个当日计划 ==========`);
        console.log(`计划信息: 机号=${{plan["飞机注册号"]}}, 出发=${{plan["出发城市"]}}, 到达=${{plan["到达城市"]}}, 日期=${{plan["出发日期"]}}`);

        let matchedRow = await tryMatchPlan(plan);
        let success = false;
        if (matchedRow) {{
            console.log(`✅ 匹配到已有航段，将尝试自动点击执行并填写`);
            success = await processExistingPlanWithManual(null, plan);
        }} else {{
            console.log(`⚠️ 未匹配到已有航段，请手动点击该计划的“执行”按钮，脚本将自动填写表单`);
            success = await processExistingPlanWithManual(null, plan);
        }}
        if (!success) {{
            console.error(`⚠️ 第 ${{processedCount}} 个计划处理失败，尝试继续下一个...`);
            await ensureListPage();
            await sleep(2000);
        }}
        await sleep(1000);
    }}
    await ensureListPage();
    console.log('🎉 所有当日计划处理完毕！');
}}
"""
    return script

def generate_nextday_script(records, city_map_json, detail_map_json, domestic_keywords_json):
    flight_records_json = json.dumps(records, ensure_ascii=False, indent=4)
    template = f"""
// ================= 自动生成的飞行计划脚本（次日计划专用） =================
// 次日待处理计划数: {len(records)}
// =========================================================

async function processNextDayRecord(record) {{
    console.log(`\\n开始处理次日计划：${{record.reg}} - ${{record.dep_city}} -> ${{record.arr_city}}`);

    if (!(await ensureListPage())) {{
        console.error('无法返回列表页，终止流程');
        return false;
    }}

    const addBtn = await waitForElement('input.query.yuanjiao', 15000, false);
    if (!addBtn) {{
        console.error('未找到添加按钮，终止流程');
        return false;
    }}
    addBtn.click();
    console.log('已点击添加按钮，等待表单加载...');

    const aircraftSelectBtn = await waitForElement('//*[@id="ele7"]', 15000, true);
    if (!aircraftSelectBtn) {{
        console.error('未找到飞机机号选择按钮，终止流程');
        return false;
    }}
    console.log('表单加载完成，找到飞机机号选择按钮');

    aircraftSelectBtn.click();
    if (!(await selectAircraft(record.reg))) {{
        console.error('选择飞机失败，终止流程');
        return false;
    }}

    const doc = await getMainDoc();
    const specialSelect = doc.querySelector('#specialf');
    if (specialSelect) await setSelectValue(specialSelect, "否");
    else console.warn('未找到是否特殊任务飞行 select');

    const certSelect = doc.querySelector('#operationCertificate');
    if (certSelect) await setSelectValue(certSelect, "是");
    else console.warn('未找到是否有运行合格证 select');

    const operateSelect = doc.querySelector('#businessOperation');
    if (operateSelect) await setSelectValue(operateSelect, "否");
    else console.warn('未找到是否经营性作业 select');

    let purpose = "自用飞行";
    const purposeRaw = record.purpose || "";
    if (purposeRaw.includes("维修") || purposeRaw.includes("调机")) {{
        purpose = "调机";
    }}
    const purposeSelect = await waitForElement('//*[contains(text(), "非经营活动")]/following-sibling::*//select', 10000, true);
    if (purposeSelect) await setSelectValue(purposeSelect, purpose);
    else console.warn('未找到用途下拉框');

    const startDateInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[9]/div/input', 5000, true);
    const endDateInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[10]/div/input', 5000, true);
    if (startDateInput) setDateInput(startDateInput, record.start_date);
    else console.warn('未找到服务开始日期输入框');
    if (endDateInput) setDateInput(endDateInput, record.end_date);
    else console.warn('未找到服务结束日期输入框');

    await fillFirstSegmentSelects(record.dep_city);
    await fillFirstSegmentTime(record);

    const addSegmentBtn = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[23]/div[1]/div[1]/div/div/button', 5000, true);
    if (addSegmentBtn) {{
        addSegmentBtn.click();
        console.log('已点击添加航段按钮，等待新航段加载...');
        await sleep(1000);
        await fillSecondSegmentSelects(record.arr_city);
        await fillSecondSegmentTime();
    }} else {{
        console.warn('未找到添加航段按钮');
    }}

    const detailAreaInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[25]/div/input', 5000, true);
    if (detailAreaInput) {{
        detailAreaInput.value = `${{record.dep_city}}-${{record.arr_city}}`;
        detailAreaInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        console.log('已填入详细作业地区');
    }} else console.warn('未找到详细作业地区输入框');

    const customerInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[27]/div/input', 5000, true);
    if (customerInput) {{
        customerInput.value = "天成商务航空有限公司";
        customerInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        console.log('已填入服务客户名称');
    }} else console.warn('未找到服务客户名称输入框');

    const baseInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[28]/div/input', 5000, true);
    if (baseInput) {{
        baseInput.value = `${{record.dep_city}}机场-${{record.arr_city}}机场`;
        baseInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        console.log('已填入作业基地名称');
    }} else console.warn('未找到作业基地名称输入框');

    const operatorInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[29]/div/input', 5000, true);
    if (operatorInput) {{
        operatorInput.value = "张永一";
        operatorInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        console.log('已填入作业负责人姓名');
    }} else console.warn('未找到作业负责人姓名输入框');

    const phoneInput = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[30]/div/input', 5000, true);
    if (phoneInput) {{
        phoneInput.value = "18566725728";
        phoneInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
        console.log('已填入负责人联系电话');
    }} else console.warn('未找到负责人联系电话输入框');

    const contractSelect = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[36]/div/select', 5000, true);
    if (contractSelect) await setSelectValue(contractSelect, "已签订");
    else console.warn('未找到合同订立情况下拉框');

    const insuranceSelect = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[37]/div/select', 5000, true);
    if (insuranceSelect) await setSelectValue(insuranceSelect, "已参保");
    else console.warn('未找到保险情况下拉框');

    // ========== 修改点：提交后等待用户手动点击确定 ==========
    const submitBtn = await waitForElement('/html/body/div[1]/div/div[3]/div/div[2]/form/div[40]/ul/li[2]/input', 5000, true);
    if (submitBtn) {{
        submitBtn.click();
        console.log('✅ 已点击“提交”按钮，弹窗已出现，请手动点击“确定”完成提交。');
        console.log('⏳ 等待30秒，以便您手动操作...');
        await sleep(30000);
        console.log('⏳ 继续执行下一个计划（假设您已手动提交完成）。');
        const backBtn = await waitForElement('input.query.yuanjiao', 5000, false);
        if (backBtn) {{
            console.log('✅ 已返回列表页');
        }} else {{
            console.warn('⚠️ 未检测到返回列表页，继续下一个计划，请确保已手动提交。');
        }}
        console.log(`处理完成：${{record.reg}}`);
    }} else {{
        console.warn('未找到提交按钮');
    }}
    await sleep(2000);
    return true;
}}

async function runNextDayPlans() {{
    const flightRecords = {flight_records_json};
    console.log(`🚀 开始执行次日计划自动化流程，共 ${{flightRecords.length}} 条计划...`);
    for (let i = 0; i < flightRecords.length; i++) {{
        const success = await processNextDayRecord(flightRecords[i]);
        if (!success) {{
            console.error(`第 ${{i+1}} 条次日计划处理失败，终止后续执行。`);
            break;
        }}
    }}
    console.log("所有次日计划处理完毕");
}}
"""
    return template

def run_feature_d():
    st.markdown("上传 Excel 文件，自动生成浏览器控制台脚本，**先自动填入当日已执飞计划，再自动备案次日计划**。")

    header_row = 1

    uploaded_file = st.file_uploader("📂 上传 Excel 文件（航段数据）", type=["xlsx", "xls"], key="d_upload")

    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file, header=header_row)
            df.columns = df.columns.str.strip()
            df = df.dropna(how='all')
            st.success("文件上传成功！")
            st.subheader("📊 数据预览（前5行）")
            st.dataframe(df.head())

            required_cols = ["飞机注册号", "出发日期", "到达日期", "用途", "出发城市", "到达城市", "预计飞行时间", "实际到达"]
            missing = [col for col in required_cols if col not in df.columns]
            if missing:
                st.error(f"❌ 缺少必要列: {missing}")
                st.info(f"实际列名: {list(df.columns)}")
                return

            df_daily = df[df["实际到达"].notna() & (df["实际到达"].astype(str).str.strip() != "")].copy()
            df['出发日期'] = pd.to_datetime(df['出发日期']).dt.date
            today = date.today()
            tomorrow = today + timedelta(days=1)
            df_nextday = df[df['出发日期'] == tomorrow].copy()
            st.info(f"✅ 共读取 {len(df)} 条飞行计划，其中当日计划（已执飞）: {len(df_daily)} 条，次日计划（出发日期为 {tomorrow}）: {len(df_nextday)} 条")

            if len(df_daily) == 0 and len(df_nextday) == 0:
                st.warning("没有需要处理的计划。")
                return

            custom_detail_map = {}
            city_map, detail_map = build_city_mappings(df, custom_detail_map)
            city_map_json = json.dumps(city_map, ensure_ascii=False, indent=4)
            detail_map_json = json.dumps(detail_map, ensure_ascii=False, indent=4)
            domestic_keywords_json = json.dumps(DOMESTIC_KEYWORDS)

            daily_records = df_daily.to_dict(orient="records")
            for rec in daily_records:
                for k, v in rec.items():
                    if pd.isna(v):
                        rec[k] = ""

            nextday_records = []
            for _, row in df_nextday.iterrows():
                purpose_raw = row.get("用途", "")
                if "维修" in purpose_raw or "调机" in purpose_raw:
                    purpose = "调机"
                else:
                    purpose = "自用飞行"
                start_date = str(row["出发日期"])
                end_date = str(row["到达日期"])
                flight_time = row.get("预计飞行时间", "")
                hours, minutes = parse_flight_time(flight_time)
                dep_city = str(row["出发城市"]).strip()
                arr_city = str(row["到达城市"]).strip()
                reg_raw = str(row["飞机注册号"]).strip()
                nextday_records.append({
                    "reg": reg_raw,
                    "start_date": start_date,
                    "end_date": end_date,
                    "purpose": purpose,
                    "dep_city": dep_city,
                    "arr_city": arr_city,
                    "flight_hours": hours,
                    "flight_minutes": minutes
                })

            base_script = generate_base_script(city_map_json, detail_map_json, domestic_keywords_json)
            daily_script = generate_daily_script(daily_records, city_map_json, detail_map_json, domestic_keywords_json) if len(daily_records) > 0 else ""
            nextday_script = generate_nextday_script(nextday_records, city_map_json, detail_map_json, domestic_keywords_json) if len(nextday_records) > 0 else ""

            final_script = base_script + "\n\n" + daily_script + "\n\n" + nextday_script + """
(async () => {
    console.log("========== 开始执行综合流程 ==========");
    if (typeof runDailyPlans === 'function') {
        await runDailyPlans();
    } else {
        console.log("没有当日计划需要处理。");
    }
    if (typeof runNextDayPlans === 'function') {
        await runNextDayPlans();
    } else {
        console.log("没有次日计划需要处理。");
    }
    console.log("========== 综合流程全部完成 ==========");
})();
"""
            st.success("脚本生成成功！")
            st.subheader("📋 复制以下代码到浏览器控制台（F12）运行")
            st.code(final_script, language="javascript")
            st.info("💡 提示：请确保已登录系统并停留在「经营活动信息管理」列表页，脚本将自动处理当日已执飞计划和次日未执飞计划。")
            st.download_button(
                label="💾 下载脚本文件 (.js)",
                data=final_script,
                file_name="flight_plan_combined.js",
                mime="application/javascript"
            )
        except Exception as e:
            st.error(f"处理文件时出错: {e}")
            import traceback
            st.exception(e)
    else:
        st.info("请上传 Excel 文件开始")
        # ==============================
# 功能 E：值班连班统计
# ==============================
def run_feature_e():
    st.markdown("上传值班表（PDF或Excel），自动统计运管主班、运控白班/夜班、补贴天数和休息天数。")

    uploaded_file = st.file_uploader("上传值班表（PDF或Excel）", type=["pdf", "xlsx", "xls"], key="e_upload")

    control_staff_input = st.text_input(
        "值班人员名单（空格分隔）",
        value="陈宇鸣 周贤民 吴迪 王浩宇 林泓辰 陈育盛 钟洪达",
        key="e_control"
    )

    if uploaded_file:
        control_staff = set(control_staff_input.strip().split())
        target_staff = control_staff

        schedules = []
        file_type = uploaded_file.type

        if file_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
            try:
                df = pd.read_excel(uploaded_file, header=None, engine='openpyxl')
            except Exception as e:
                st.error(f"读取Excel失败: {e}")
                st.stop()

            st.subheader("原始表格预览（前20行）")
            st.dataframe(df.head(20))

            header_idx = None
            for i, row in df.iterrows():
                row_str = " ".join([str(c) for c in row if pd.notna(c)])
                if "运行管理" in row_str:
                    header_idx = i
                    break

            if header_idx is None:
                st.error("未找到包含'运行管理'的表头行")
                st.stop()

            col_mapping = {}
            for idx, val in enumerate(df.iloc[header_idx]):
                val_str = str(val) if pd.notna(val) else ""
                if "运行管理" in val_str:
                    col_mapping[idx] = "management"
                elif "运行计划" in val_str:
                    col_mapping[idx] = "plan"
                elif "运行监控" in val_str:
                    col_mapping[idx] = "control"
                elif "运行保障" in val_str:
                    col_mapping[idx] = "control"
                elif "运行支援" in val_str:
                    col_mapping[idx] = "control"

            data_start = header_idx + 1
            current_date = None
            day_row = None

            for i in range(data_start, len(df)):
                row = df.iloc[i]
                second_cell = str(row[1]) if pd.notna(row[1]) else ""
                second_cell = second_cell.strip()

                if "白" in second_cell:
                    first_cell = str(row[0]) if pd.notna(row[0]) else ""
                    date_match = re.search(r"(\d+月\d+日|\d+日)", first_cell)
                    if date_match:
                        current_date = date_match.group(1)
                    else:
                        current_date = first_cell
                    day_row = row
                elif "晚" in second_cell and day_row is not None:
                    if current_date is None:
                        first_cell = str(row[0]) if pd.notna(row[0]) else ""
                        date_match = re.search(r"(\d+月\d+日|\d+日)", first_cell)
                        if date_match:
                            current_date = date_match.group(1)
                    if current_date:
                        day_people = set()
                        night_people = set()
                        for col_idx, role in col_mapping.items():
                            day_name = str(day_row[col_idx]).strip() if pd.notna(day_row[col_idx]) else ""
                            night_name = str(row[col_idx]).strip() if pd.notna(row[col_idx]) else ""
                            if day_name and day_name not in ["nan", "None", ""]:
                                if day_name in target_staff:
                                    day_people.add(day_name)
                            if night_name and night_name not in ["nan", "None", ""]:
                                if night_name in target_staff:
                                    night_people.add(night_name)
                        schedules.append({"date": current_date, "day": day_people, "night": night_people})
                        day_row = None
                        current_date = None

            if not schedules:
                st.error("未能从Excel解析到排班数据，请检查格式")
                st.stop()

        else:
            with pdfplumber.open(uploaded_file) as pdf:
                all_text = ""
                for page in pdf.pages:
                    all_text += page.extract_text() + "\n"

            lines = all_text.split("\n")
            day_shifts = []
            night_shifts = []

            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if "白" in line and "晚" not in line:
                    date_match = re.search(r"(\d+月\d+日|\d+日)", line)
                    date_str = date_match.group(1) if date_match else ""
                    names = re.findall(r"[\u4e00-\u9fa5]{2,3}", line)
                    filtered_names = [n for n in names if n in target_staff and n not in ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日", "运行控制", "运行管理", "白班", "夜班", "带班主任"]]
                    if filtered_names:
                        day_shifts.append((date_str, filtered_names))
                elif "晚" in line:
                    date_match = re.search(r"(\d+月\d+日|\d+日)", line)
                    date_str = date_match.group(1) if date_match else ""
                    names = re.findall(r"[\u4e00-\u9fa5]{2,3}", line)
                    filtered_names = [n for n in names if n in target_staff and n not in ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日", "运行控制", "运行管理", "白班", "夜班", "带班主任"]]
                    if filtered_names:
                        night_shifts.append((date_str, filtered_names))

            min_len = min(len(day_shifts), len(night_shifts))
            for i in range(min_len):
                date_day, day_names = day_shifts[i]
                date_night, night_names = night_shifts[i]
                date_str = date_day if date_day else date_night
                schedules.append({
                    "date": date_str,
                    "day": set(day_names),
                    "night": set(night_names)
                })

            if not schedules:
                st.error("未识别到任何排班数据，请检查文件格式")
                st.stop()

        st.success(f"成功识别 {len(schedules)} 天的排班数据")

        import re as _re
        def parse_date_to_tuple(date_str):
            match = _re.search(r'(\d+)月(\d+)日', date_str)
            if match:
                return (int(match.group(1)), int(match.group(2)))
            match = _re.search(r'(\d+)日', date_str)
            if match:
                return (9, int(match.group(1)))
            return (9, 99)

        schedules.sort(key=lambda x: parse_date_to_tuple(x['date']))

        with st.expander("🔍 完整排班数据（请核对）"):
            st.write(f"**总天数：{len(schedules)} 天**")
            for sch in schedules:
                st.write(f"**{sch['date']}**")
                st.write(f"  - 白班：{', '.join(sorted(sch['day'])) if sch['day'] else '（无）'}")
                st.write(f"  - 夜班：{', '.join(sorted(sch['night'])) if sch['night'] else '（无）'}")

        all_persons = target_staff
        stats = {name: {"consecutive": 0, "pure_day": 0, "pure_night": 0, "total_night": 0, "rest_days": 0} for name in all_persons}
        attendance_records = {name: [] for name in all_persons}

        for sch in schedules:
            date_str = sch["date"]
            day_set = sch["day"]
            night_set = sch["night"]

            for name in all_persons:
                in_day = name in day_set
                in_night = name in night_set
                attendance_records[name].append(in_day or in_night)

                if in_day and in_night:
                    stats[name]["consecutive"] += 1
                elif in_day and not in_night:
                    stats[name]["pure_day"] += 1
                elif not in_day and in_night:
                    stats[name]["pure_night"] += 1

        for name in all_persons:
            stats[name]["total_night"] = stats[name]["consecutive"] + stats[name]["pure_night"]

        for name in all_persons:
            rest = 0
            cnt = 0
            for present in attendance_records[name]:
                if not present:
                    cnt += 1
                else:
                    if cnt >= 2:
                        rest += (cnt - 1)
                    cnt = 0
            if cnt >= 2:
                rest += (cnt - 1)
            stats[name]["rest_days"] = rest

        result_data = []
        for name in all_persons:
            consecutive = stats[name]["consecutive"]
            pure_day = stats[name]["pure_day"]
            pure_night = stats[name]["pure_night"]

            main_shift_min = 15 * 60 + 30
            day_shift_min = 8 * 60 + 30
            night_shift_min = 15 * 60 + 30

            total_minutes = consecutive * main_shift_min + pure_day * day_shift_min + pure_night * night_shift_min
            hours = total_minutes // 60
            minutes = total_minutes % 60
            total_time_str = f"{hours}:{minutes:02d}"

            result_data.append({
                "姓名": name,
                "运管主班": consecutive,
                "运控白班": pure_day,
                "运控夜班": pure_night,
                "补贴天数": stats[name]["total_night"],
                "休息天数": stats[name]["rest_days"],
                "累计在岗时间": total_time_str
            })

        result_df = pd.DataFrame(result_data).sort_values(by="运管主班", ascending=False)

        st.subheader("📌 值班人员")
        st.dataframe(result_df, use_container_width=True, height=400)

        csv = result_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button("📥 下载完整统计表 (CSV)", csv, "shift_statistics.csv", "text/csv", key="e_download")

# ==============================
# 功能 F：世界时行程
# ==============================
def run_feature_f():
    st.markdown("从Jetops系统导出的北京时间的行程 Excel 文件转换为世界时的行程，便于复制粘贴。")
    st.info("💡 每次上传将自动与上一次记录对比，新增或变更的航段会在下方红色高亮显示。")

    HISTORY_FILE = "flight_history.json"

    def load_history():
        if os.path.exists(HISTORY_FILE):
            try:
                with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                return {"records": []}
        else:
            return {"records": []}

    def save_history(history):
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)

    def parse_time_column(val):
        if pd.isna(val):
            return None
        if isinstance(val, (pd.Timestamp, datetime)):
            return val.strftime('%H:%M')
        if hasattr(val, 'strftime'):
            return val.strftime('%H:%M')
        s = str(val).strip()
        if ':' in s:
            return s[:5]
        return s

    def convert_to_utc(date_val, time_str):
        if pd.isna(date_val) or time_str is None:
            return None
        if isinstance(date_val, (pd.Timestamp, datetime)):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val).split()[0]
        dt_str = f"{date_str} {time_str}"
        try:
            dt_local = pd.to_datetime(dt_str)
            dt_utc = dt_local - timedelta(hours=8)
            return dt_utc
        except:
            return None

    def format_utc(dt):
        if dt is None:
            return ""
        months = ['JAN','FEB','MAR','APR','MAY','JUN',
                  'JUL','AUG','SEP','OCT','NOV','DEC']
        day = dt.day
        month = months[dt.month - 1]
        hour = dt.hour
        minute = dt.minute
        return f"{day:02d}{month} {hour:02d}{minute:02d}Z"

    def generate_plans(df):
        required = ['飞机注册号', '出发地', '到达地', '出发日期', '计划出发', '到达日期', '预计到达', '用途']
        for col in required:
            if col not in df.columns:
                st.error(f"❌ 缺少列：{col}")
                return None

        df = df.dropna(subset=['出发地', '到达地', '出发日期', '计划出发'])
        if df.empty:
            st.warning("没有有效的航段数据")
            return None

        plans = {}
        for idx, row in df.iterrows():
            reg = row['飞机注册号']
            if pd.isna(reg) or str(reg).strip() == '':
                reg = "N/A"
            else:
                reg = str(reg).strip()

            dep_time = parse_time_column(row['计划出发'])
            arr_time = parse_time_column(row['预计到达'])
            if dep_time is None or arr_time is None:
                continue

            dep_utc = convert_to_utc(row['出发日期'], dep_time)
            arr_utc = convert_to_utc(row['到达日期'], arr_time)
            if dep_utc is None or arr_utc is None:
                continue

            use = str(row['用途']) if not pd.isna(row['用途']) else ''
            flight_type = 'FERRY' if '调机' in use else 'PAX'

            line = (f"ETD {row['出发地']} {format_utc(dep_utc)} // "
                    f"ETA {row['到达地']} {format_utc(arr_utc)}  {flight_type}")

            if reg not in plans:
                plans[reg] = []
            plans[reg].append((dep_utc, line))

        result = {}
        for reg, items in plans.items():
            items.sort(key=lambda x: x[0])
            lines = [reg]
            lines.extend([item[1] for item in items])
            result[reg] = "\n".join(lines)

        return result

    def sort_plans(plans_dict):
        priority_order = ['B652Q', 'B65AP', 'B652S', 'MLLIN', 'N88AY', 'B652R']
        all_keys = list(plans_dict.keys())
        priority_keys = [k for k in priority_order if k in all_keys]
        remaining_keys = [k for k in all_keys if k not in priority_order and k != "N/A"]
        remaining_keys.sort()
        na_keys = [k for k in all_keys if k == "N/A"]
        sorted_keys = priority_keys + remaining_keys + na_keys
        return {k: plans_dict[k] for k in sorted_keys}

    def diff_plans(old_plans, new_plans):
        changes = {}
        all_regs = set(old_plans.keys()) | set(new_plans.keys())
        for reg in all_regs:
            old_lines = set(old_plans.get(reg, "").split('\n')) if old_plans.get(reg) else set()
            new_lines = set(new_plans.get(reg, "").split('\n')) if new_plans.get(reg) else set()
            old_lines.discard(reg)
            new_lines.discard(reg)
            added = new_lines - old_lines
            for line in added:
                changes[(reg, line)] = 'added'
        return changes

    uploaded_file_2 = st.file_uploader("📤 上传航段数据导出（北京时间）", type=["xlsx"], key="f_worldtime")

    if uploaded_file_2 is not None:
        try:
            df = pd.read_excel(uploaded_file_2, skiprows=1)
            st.success("✅ 文件读取成功")

            new_plans = generate_plans(df)
            if new_plans is None:
                st.stop()

            sorted_new_plans = sort_plans(new_plans)

            history = load_history()
            old_plans = {}
            if history["records"]:
                last_record = history["records"][-1]
                old_plans = last_record.get("data", {})

            changes = diff_plans(old_plans, new_plans)

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            filename = uploaded_file_2.name
            new_record = {
                "timestamp": timestamp,
                "filename": filename,
                "data": new_plans
            }
            history["records"].append(new_record)
            if len(history["records"]) > 20:
                history["records"] = history["records"][-20:]
            save_history(history)

            st.subheader("📋 生成的飞行计划（红色为新增/变更）")

            for reg, text in sorted_new_plans.items():
                lines = text.split('\n')
                has_changes = any((reg, line) in changes for line in lines if line != reg)

                if has_changes:
                    st.markdown(f"**✈️ {reg}** 🔴 <span style='color:red;font-size:0.9rem;'>（有新增或变更）</span>", unsafe_allow_html=True)
                else:
                    st.markdown(f"**✈️ {reg}**")

                plain_lines = []
                for line in lines:
                    if line == reg:
                        continue
                    plain_lines.append(line)

                plain_text = "\n".join(plain_lines)
                st.code(plain_text, language="text")

            full_text = ""
            for reg, text in sorted_new_plans.items():
                full_text += f"{text}\n\n"
            with st.expander("📦 全部计划合并（点击展开）"):
                st.code(full_text, language="text")

            with st.expander("📜 查看历史上传记录"):
                if history["records"]:
                    for i, rec in enumerate(history["records"]):
                        st.write(f"{i+1}. {rec['timestamp']} - {rec['filename']}")
                else:
                    st.write("暂无历史记录")

            if st.button("🗑️ 清除所有历史记录", key="clear_history_f"):
                save_history({"records": []})
                st.success("历史已清除，请刷新页面")
                st.rerun()

        except Exception as e:
            st.error(f"❌ 处理出错：{e}")
            st.stop()
    else:
        st.info("请上传一个符合格式的 Excel 文件。")

    st.markdown("---")
    st.caption("🛠️ 工具说明：日期/时间按北京时间（UTC+8）自动转换为世界时（Z）。对比功能基于上一次上传的记录。")

# ==============================
# 功能 G：航路处理工具
# ==============================
def run_feature_g():
    st.markdown("支持表格格式（带N/E坐标）/中文描述格式，自动精简航路+添加#前缀，兼容不规整数据")

    def parse_coord(coord_str):
        letter = coord_str[0]
        num_part = coord_str[1:]
        if letter == 'N':
            deg = int(num_part[0:2])
            minute = int(num_part[2:4])
            sec_part = num_part[4:]
            if '.' in sec_part:
                sec_float = float(sec_part)
                sec_int = int(round(sec_float))
            else:
                sec_int = int(sec_part)
            if sec_int >= 60:
                sec_int -= 60
                minute += 1
                if minute >= 60:
                    minute -= 60
                    deg += 1
            return f"{deg:02d}{minute:02d}{sec_int:02d}"
        elif letter == 'E':
            deg = int(num_part[0:3])
            minute = int(num_part[3:5])
            sec_part = num_part[5:]
            if '.' in sec_part:
                sec_float = float(sec_part)
                sec_int = int(round(sec_float))
            else:
                sec_int = int(sec_part)
            if sec_int >= 60:
                sec_int -= 60
                minute += 1
                if minute >= 60:
                    minute -= 60
                    deg += 1
            return f"{deg:03d}{minute:02d}{sec_int:02d}"
        else:
            raise ValueError(f"未知的坐标前缀: {letter}")

    def base_name(s):
        return s.split('@')[0]

    def is_open_point(s):
        base = base_name(s)
        if re.match(r'^[A-Z]{2,5}$', base):
            return True
        if re.match(r'^P[A-Z]+$', base):
            return True
        return False

    def is_p_point(s):
        base = base_name(s)
        return re.match(r'^P\d+$', base) is not None

    def clean_route(r):
        if r.startswith('#'):
            return r[1:]
        return r

    def is_open_route(rt):
        return rt and rt[0] not in ('H', 'J', 'V')

    def extract_table(text):
        tokens = text.strip().split()
        start_idx = 0
        for i, tok in enumerate(tokens):
            if tok.isdigit() and 1 <= int(tok) <= 40:
                start_idx = i
                break
        tokens = tokens[start_idx:]

        lines = []
        i = 0
        while i < len(tokens):
            if tokens[i].isdigit():
                line = [tokens[i]]
                i += 1
                while i < len(tokens) and not tokens[i].isdigit():
                    line.append(tokens[i])
                    i += 1
                lines.append(line)

        points = []
        routes = []
        for line in lines:
            lat_idx = None
            for idx, tok in enumerate(line):
                if tok.startswith('N') and tok[1:].replace('.', '', 1).isdigit():
                    lat_idx = idx
                    break
            if lat_idx is None:
                continue
            lon_idx = lat_idx + 1
            if lon_idx >= len(line) or not line[lon_idx].startswith('E'):
                continue
            lat_str = line[lat_idx]
            lon_str = line[lon_idx]

            route = None
            if lon_idx + 1 < len(line):
                next_tok = line[lon_idx + 1]
                if re.match(r'^[A-Z][A-Z0-9]*$', next_tok) and not next_tok[0].isdigit():
                    route = next_tok

            point_name = None
            for j in range(lat_idx - 1, 0, -1):
                tok = line[j]
                if is_open_point(tok) or is_p_point(tok):
                    point_name = tok
                    break
            if point_name is None:
                continue

            if is_p_point(point_name):
                lat_int = parse_coord(lat_str)
                lon_int = parse_coord(lon_str)
                point_display = f"{point_name}@{lat_int}N{lon_int}E"
            else:
                point_display = point_name

            points.append(point_display)
            if route is not None:
                routes.append(route)

        seq = []
        for i in range(len(points)):
            seq.append(points[i])
            if i < len(routes):
                seq.append(routes[i])
        return seq

    def extract_chinese(text):
        text = re.sub(r'[\u4e00-\u9fa5，、。；：""''（）【】]', ' ', text)
        words = text.split()
        seq = []
        for w in words:
            if '(' in w and ')' in w:
                m = re.search(r'\(([A-Z]+)\)', w)
                if m:
                    point = m.group(1)
                    prefix = w[:w.find('(')]
                    m_route = re.search(r'([A-Z]\d+)$', prefix)
                    if m_route:
                        seq.append(m_route.group(1))
                    seq.append(point)
            elif re.match(r'^[A-Z]\d+[A-Z]{2,5}$', w) or re.match(r'^[A-Z]\d+P\d+$', w):
                m = re.match(r'^([A-Z]\d+)([A-Z]{2,5}|P\d+)$', w)
                if m:
                    seq.append(m.group(1))
                    seq.append(m.group(2))
            elif re.match(r'^[A-Z]\d+$', w):
                seq.append(w)
            elif is_open_point(w) or is_p_point(w):
                seq.append(w)
        return seq

    def step1_extract(text):
        if re.search(r'N\d{5,6}(?:\.\d+)?\s+E\d{6,7}(?:\.\d+)?', text):
            return extract_table(text), 'table'
        else:
            return extract_chinese(text), 'chinese'

    def step2_reduce(seq):
        L = seq[:]
        changed = True
        while changed:
            changed = False
            n = len(L)
            candidates = []
            for i in range(0, n, 2):
                if not is_open_point(L[i]):
                    continue
                if i + 1 >= n:
                    continue
                first_route = clean_route(L[i+1])
                if not is_open_route(first_route):
                    continue
                for j in range(i+2, n, 2):
                    all_same = True
                    for k in range(i+1, j, 2):
                        rt = clean_route(L[k])
                        if rt != first_route or not is_open_route(rt):
                            all_same = False
                            break
                    if not all_same:
                        break
                    if is_open_point(L[j]):
                        length = (j - i) // 2
                        if length >= 2:
                            candidates.append((i, j, length))
            if not candidates:
                break
            candidates.sort(key=lambda x: -x[2])
            best_i, best_j, _ = candidates[0]
            new_segment = [L[best_i], L[best_i+1], L[best_j]]
            L = L[:best_i] + new_segment + L[best_j+1:]
            changed = True
        return L

    def step3_add_hash(seq):
        pts = seq[0::2]
        rts = seq[1::2]
        m = len(rts)

        def is_closed_route(rt):
            return rt.startswith(('H', 'J', 'V'))

        def is_p(pt):
            base = base_name(pt)
            return re.match(r'^P\d+$', base) is not None

        res = [pts[0]]
        for i, rt in enumerate(rts):
            left = pts[i]
            right = pts[i+1]
            need_hash = False
            if is_closed_route(rt):
                need_hash = True
            elif is_p(left) or is_p(right):
                need_hash = True
            res.append('#' + rt if need_hash else rt)
            res.append(right)
        return res

    if "last_processed_input_route" not in st.session_state:
        st.session_state.last_processed_input_route = ""
    if "result_text_route" not in st.session_state:
        st.session_state.result_text_route = ""

    input_text_route = st.text_area(
        "📋 请输入待处理的航路文本",
        key="input_text_route_g",
        height=300,
        placeholder="粘贴民航航线数据，支持多行表格格式/纯中文描述格式..."
    )

    btn_col1, btn_col2, btn_col3 = st.columns([2, 2, 8])
    with btn_col1:
        process_btn = st.button("⚙️ 处理", type="primary", use_container_width=True, key="process_route_g")
    with btn_col2:
        clear_btn = st.button("🗑️ 清空", use_container_width=True, key="clear_route_g")

    if clear_btn:
        st.session_state.input_text_route_g = ""
        st.session_state.last_processed_input_route = ""
        st.session_state.result_text_route = ""
        st.rerun()

    if process_btn and st.session_state.get("input_text_route_g", "").strip():
        progress_bar = st.progress(0)
        status_text = st.empty()
        total_steps = 4
        current_step = 0

        try:
            current_step += 1
            progress_bar.progress(current_step / total_steps)
            status_text.text(f"处理中：第{current_step}步/共{total_steps}步（识别输入类型）")
            seq, fmt = step1_extract(st.session_state.input_text_route_g)

            current_step += 1
            progress_bar.progress(current_step / total_steps)
            status_text.text(f"处理中：第{current_step}步/共{total_steps}步（精简相同开放航路）")
            if fmt == 'table':
                seq = step2_reduce(seq)

            current_step += 1
            progress_bar.progress(current_step / total_steps)
            status_text.text(f"处理中：第{current_step}步/共{total_steps}步（添加航路#前缀）")
            if fmt == 'table':
                seq = step3_add_hash(seq)

            current_step += 1
            progress_bar.progress(current_step / total_steps)
            status_text.text(f"处理中：第{current_step}步/共{total_steps}步（生成最终结果）")
            result = ' '.join(seq) if seq else "⚠️ 未提取到有效航路数据"

            st.session_state.result_text_route = result
            st.session_state.last_processed_input_route = st.session_state.input_text_route_g

            progress_bar.empty()
            status_text.empty()
            st.success("✅ 处理完成！结果如下：")

        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            st.error(f"❌ 处理失败：{str(e)}")
            with st.expander("🔍 查看详细错误信息", expanded=False):
                st.code(traceback.format_exc(), language="text")

    if st.session_state.get("result_text_route", ""):
        current_input = st.session_state.get("input_text_route_g", "")
        last_input = st.session_state.last_processed_input_route

        st.subheader("📊 处理结果", divider="blue")

        if current_input != last_input:
            st.warning("⚠️ 输入已更改，当前显示的是上一次处理的结果，如需更新请点击「处理」按钮。")

        st.code(st.session_state.result_text_route, language="text")

    if not st.session_state.get("result_text_route", "") and not st.session_state.get("input_text_route_g", "").strip():
        st.info("💡 提示：粘贴航路数据后，点击「处理」即可，支持30+行不规整表格数据")

    st.markdown("---")
    st.caption("✈️ 支持表格格式（带N/E坐标）/中文描述格式，自动精简航路+添加#前缀")

# ==============================
# 主界面
# ==============================
st.set_page_config(page_title="监管局的表，发个那三位", layout="wide")
st.title("监管局的表，发个那三位")

tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "📋 每日飞行数据-10：00发",
    "📄 每日通航运行情况跟踪表-16：30发",
    "📊 天成商务航空每日运行跟踪-16：30发",
    "📜 通航脚本",
    "📊 值班连班统计",
    "🌐 世界时行程",
    "✈️ 航路处理"
])

with tab1:
    run_feature_a()
with tab2:
    run_feature_b()
with tab3:
    run_feature_c()
with tab4:
    run_feature_d()
with tab5:
    run_feature_e()
with tab6:
    run_feature_f()
with tab7:
    run_feature_g()

st.caption("💡 功能1自动更新模板的汇总数据；功能2按航段逐行填写备案表；功能3生成每日运行跟踪汇总一行；功能4生成浏览器控制台脚本；功能5统计值班连班数据；功能6将北京时间行程转换为世界时；功能7航路处理工具。")
